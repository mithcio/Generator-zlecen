"""Jednorazowy import danych z plików źródłowych (źródła/) do app/data/*.json.

Uruchamiany ręcznie przy aktualizacji danych rozliczeniowych/słownikowych.
Aplikacja (app/) nigdy nie otwiera plików xlsm źródłowych w runtime.

Użycie:
    python scripts/export_seed_data.py
"""
import json
import re
import shutil
import sys
from pathlib import Path

import openpyxl
from openpyxl.formatting.formatting import ConditionalFormattingList

# Domyślny codepage konsoli Windows (cp1252) nie zna polskich znaków - bez
# tego print() z "ń"/"ó" itp. wywala UnicodeEncodeError w połowie skryptu,
# po zapisaniu części plików JSON, ale bez pozostałych (potwierdzone: właśnie
# tak się stało przy exporcie klienci_agencyjni.json).
if sys.stdout and sys.stdout.encoding and sys.stdout.encoding.lower() != "utf-8":
    sys.stdout.reconfigure(encoding="utf-8")

ROOT = Path(__file__).resolve().parent.parent
ZRODLA = ROOT / "źródła"
DATA_OUT = ROOT / "app" / "data"

GENERATOR_XLSM = ZRODLA / "Generator_zleceń_i_danych_25022026.xlsm"
KAMPANIE_XLSM = ZRODLA / "Igor_Kampanie_MediaFarm_2026.xlsm"
NUMERY_XLSX = ZRODLA / "Numery_zlecen_2026.xlsx"

# Jedna zakładka per account manager w Numery_zlecen_2026.xlsx - każda ma dwie
# tabele Excela: "Podmioty" (A:E - Nazwa/Podmiot/Adres/Numery/Termin) i
# "Klienci pod agencjami" (G:H, + I - Klient/Agencja/Termin płatności klienta,
# patrz export_terminy_platnosci_klientow). Wiersz 1 = baner instrukcji,
# wiersz 2 = nagłówki, dane od wiersza 3.
AKANCI_ARKUSZE = ["Agnieszka Kraińska", "Marta Urbańska", "Igor Samul"]


def export_podmioty(numery_xlsx=None):
    """Podmioty: per account manager -> per Agencja/Klient bezpośredni -> dane
    fakturowe, z tabeli "Podmioty" w Numery_zlecen_2026.xlsx (kolumna "Podmiot"
    = Sp. k. dla agencji, Sp. z o.o. dla klientów bezpośrednich).

    numery_xlsx: ścieżka ustawiona w panelu Ustawienia (app.services.ustawienia)
    - jeśli nieustawiona, spada na kopię w źródła/ (wygoda przy pierwszym
    uruchomieniu, zanim ktoś skonfiguruje właściwą ścieżkę)."""
    wb = openpyxl.load_workbook(numery_xlsx or NUMERY_XLSX, data_only=True)

    wynik = {}
    for akant in AKANCI_ARKUSZE:
        ws = wb[akant]
        entity_map = {}
        r = 3
        while True:
            nazwa = ws.cell(row=r, column=1).value
            if not nazwa or not str(nazwa).strip():
                break
            entity_map[str(nazwa).strip()] = {
                "adres_fakturowy": ws.cell(row=r, column=3).value,
                "numery_rejestrowe": ws.cell(row=r, column=4).value,
                "termin_platnosci": ws.cell(row=r, column=5).value,
                "domyslny_podmiot": ws.cell(row=r, column=2).value,
            }
            r += 1
        wynik[akant] = entity_map

    with open(DATA_OUT / "podmioty.json", "w", encoding="utf-8") as f:
        json.dump(wynik, f, ensure_ascii=False, indent=2)
    print(f"podmioty.json: {sum(len(v) for v in wynik.values())} podmiotów w {len(wynik)} blokach account")


def export_klienci_agencyjni(numery_xlsx=None):
    """Mapowanie klient (marka) -> agencja per account manager, z tabeli
    "Klienci pod agencjami" w Numery_zlecen_2026.xlsx. Klienci bez jeszcze
    przypisanej agencji (kolumna "Agencja" pusta) są pomijani - mapowanie jest
    budowane ręcznie i stopniowo, nie musi być kompletne od razu."""
    wb = openpyxl.load_workbook(numery_xlsx or NUMERY_XLSX, data_only=True)

    wynik = {}
    for akant in AKANCI_ARKUSZE:
        ws = wb[akant]
        mapa = {}
        r = 3
        while True:
            klient = ws.cell(row=r, column=7).value
            if not klient or not str(klient).strip():
                break
            agencja = ws.cell(row=r, column=8).value
            if agencja and str(agencja).strip():
                mapa[str(klient).strip()] = str(agencja).strip()
            r += 1
        wynik[akant] = mapa

    with open(DATA_OUT / "klienci_agencyjni.json", "w", encoding="utf-8") as f:
        json.dump(wynik, f, ensure_ascii=False, indent=2)
    print(f"klienci_agencyjni.json: {sum(len(v) for v in wynik.values())} przypisań w {len(wynik)} blokach account")


def export_terminy_platnosci_klientow(numery_xlsx=None):
    """Termin płatności przypisany wprost do klienta (kolumna I tabeli
    "Klienci pod agencjami" w Numery_zlecen_2026.xlsx) - nadrzędny względem
    terminu przypisanego do domu mediowego/agencji (kolumna E tabeli
    "Podmioty", patrz export_podmioty). Rzadki wyjątek (na razie tylko jeden
    klient u Marty Urbańskiej) - większość klientów nie ma tu nic wpisanego,
    więc mapowanie jest z założenia niepełne/rzadkie."""
    wb = openpyxl.load_workbook(numery_xlsx or NUMERY_XLSX, data_only=True)

    wynik = {}
    for akant in AKANCI_ARKUSZE:
        ws = wb[akant]
        mapa = {}
        r = 3
        while True:
            klient = ws.cell(row=r, column=7).value
            if not klient or not str(klient).strip():
                break
            termin = ws.cell(row=r, column=9).value
            if termin and str(termin).strip():
                mapa[str(klient).strip()] = str(termin).strip()
            r += 1
        wynik[akant] = mapa

    with open(DATA_OUT / "terminy_platnosci_klientow.json", "w", encoding="utf-8") as f:
        json.dump(wynik, f, ensure_ascii=False, indent=2)
    print(
        f"terminy_platnosci_klientow.json: {sum(len(v) for v in wynik.values())} nadpisań "
        f"w {len(wynik)} blokach account"
    )


def export_mediafarm():
    """Dane spółek Mediafarm (Sp. k. / Sp. z o.o.) + kontakty accountów."""
    wb = openpyxl.load_workbook(GENERATOR_XLSM, data_only=True)
    ws = wb["Podmioty"]

    mf_row = None
    for r in range(1, ws.max_row + 1):
        if ws.cell(row=r, column=2).value == "MF DANE":
            mf_row = r
            break
    if mf_row is None:
        raise RuntimeError("Nie znaleziono wiersza 'MF DANE' w arkuszu Podmioty")

    spolki = {
        "Sp. k.": {
            "nazwa": ws.cell(row=mf_row, column=3).value,
            "numery_rejestrowe": ws.cell(row=mf_row, column=6).value,
            "konto_bankowe": ws.cell(row=mf_row, column=13).value,
        },
        "Sp. z o.o.": {
            "nazwa": ws.cell(row=mf_row + 1, column=3).value,
            "numery_rejestrowe": ws.cell(row=mf_row + 1, column=6).value,
            "konto_bankowe": ws.cell(row=mf_row + 1, column=13).value,
        },
        "adres": ws.cell(row=mf_row + 2, column=3).value,
    }

    accounts_row = None
    for r in range(1, ws.max_row + 1):
        if ws.cell(row=r, column=1).value == "Akanci":
            accounts_row = r
            break
    if accounts_row is None:
        raise RuntimeError("Nie znaleziono wiersza 'Akanci' w arkuszu Podmioty")

    # Numer telefonu w arkuszu źródłowym dla Marty Urbańskiej jest nieaktualny
    # (690987004) - potwierdzone poprawne 730010094, poprawiane tu, żeby
    # ponowne uruchomienie tego skryptu nie przywróciło starego numeru.
    POPRAWKI_TELEFONOW = {"Marta Urbańska": "730010094"}

    accounts = {}
    r = accounts_row + 1
    while ws.cell(row=r, column=1).value:
        nazwa = str(ws.cell(row=r, column=1).value).strip()
        telefon = ws.cell(row=r, column=3).value
        # Surowe cyfry bez spacji/prefiksu - formatowanie ("+48 XXX XXX XXX")
        # dopiero przy wyświetlaniu, patrz lookup_podmiotu.formatuj_telefon().
        telefon_cyfry = re.sub(r"\D", "", str(telefon)) if telefon else None
        accounts[nazwa] = {
            "email": ws.cell(row=r, column=2).value,
            "telefon": POPRAWKI_TELEFONOW.get(nazwa, telefon_cyfry),
        }
        r += 1

    wynik = {"spolki": spolki, "accounts": accounts}
    with open(DATA_OUT / "mediafarm.json", "w", encoding="utf-8") as f:
        json.dump(wynik, f, ensure_ascii=False, indent=2)
    print(f"mediafarm.json: 2 spółki, {len(accounts)} accountów")


SZABLONY_OUT = DATA_OUT / "szablony_wydawcow"
ARKUSZE_WYDAWCOW = ["KIDOZ", "Adverty", "Odeeo"]


def export_szablony_wydawcow():
    """Wyodrębnia zakładki KIDOZ/Adverty/Odeeo z Generator_zleceń_i_danych.xlsm
    do osobnych plików xlsx (jeden arkusz = jeden plik), zachowując 1:1
    formatowanie (kolory, scalenia, szerokości kolumn) - to są "szablony",
    które generator_wydawcy.py kopiuje i wypełnia danymi konkretnego zlecenia
    w runtime, zamiast odtwarzać formatowanie ręcznie w kodzie. Formuły
    odwołujące się do innych zakładek (Kampania!/Zlecenie!/Podmioty!) zostają
    w szablonie tak jak są - i tak są podmieniane na wartości literalne
    dopiero przy generowaniu konkretnego pliku, bo po usunięciu pozostałych
    zakładek i tak przestałyby działać (zamieniłyby się na #REF!).

    Formatowanie warunkowe (cała zakładka na czerwono, gdy formuła stawki nie
    dopasuje żadnego znanego formatu) jest tu celowo usuwane - opierało się
    na tym, że niedopasowana formuła zwraca FALSE, a Excel traktuje 0=FALSE
    jako prawdę, więc przy naszej cenie=0 (patrz generator_wydawcy.py)
    zapalałoby się zawsze, mimo że to nie błąd. Poprawność pól pilnujemy
    teraz sami w kodzie, nie formułą w arkuszu.

    Zewnętrzny link do Igor_Kampanie_MediaFarm_2026.xlsm ([1]Dane, używany
    tylko w formułach zakładki Kampania - patrz P5 w tamtym arkuszu) też jest
    usuwany - żaden z arkuszy KIDOZ/Adverty/Odeeo go nie używa, a zostawiony
    jako osierocone odwołanie po usunięciu Kampanii Excel zgłaszał przy
    otwieraniu wygenerowanego pliku jako uszkodzoną zawartość.

    Podobnie usuwane: workbookPr/codeName="ThisWorkbook" i sheetPr/codeName
    zakładki (wiążą się z modułami VBA - mimo braku samego kodu makra te
    atrybuty same w sobie sprawiały, że Excel zgłaszał plik jako "macro-free,
    ale zawiera zawartość makro-enabled") oraz zdefiniowane nazwy
    "_xleta.IF"/"_xleta.TODAY" (makra Excel 4.0/XLM, oznaczone xlm="1", obie
    i tak zepsute - #NAME?)."""
    SZABLONY_OUT.mkdir(parents=True, exist_ok=True)
    for nazwa_arkusza in ARKUSZE_WYDAWCOW:
        wb = openpyxl.load_workbook(GENERATOR_XLSM, keep_vba=False, data_only=False)
        for inny in list(wb.sheetnames):
            if inny != nazwa_arkusza:
                del wb[inny]
        wb._external_links = []
        wb.code_name = None
        wb.defined_names.clear()
        ws = wb.active
        ws.sheet_properties.codeName = None
        ws.conditional_formatting = ConditionalFormattingList()
        plik = SZABLONY_OUT / f"{nazwa_arkusza.lower()}.xlsx"
        wb.save(plik)
        print(f"{plik.name}: wyodrębniono zakładkę '{nazwa_arkusza}' (oczyszczono z formatowania warunkowego, external linku i śladów VBA/XLM)")


SZABLON_CRAZYGAMES_ZRODLO = ZRODLA / "Crazygames_Mediafarm_UIP_Pan_Wilk_IO.xlsx"


def export_szablon_crazygames():
    """Kopiuje przykładowy plik IO dla Crazygames jako szablon 1:1 - w
    odróżnieniu od KIDOZ/Adverty/Odeeo to już gotowy, prosty plik xlsx (nie
    wycinek z dużego xlsm), więc wystarczy kopia, bez czyszczenia."""
    if not SZABLON_CRAZYGAMES_ZRODLO.exists():
        print(f"Pominięto szablon Crazygames - brak pliku {SZABLON_CRAZYGAMES_ZRODLO}")
        return
    SZABLONY_OUT.mkdir(parents=True, exist_ok=True)
    plik = SZABLONY_OUT / "crazygames.xlsx"
    shutil.copyfile(SZABLON_CRAZYGAMES_ZRODLO, plik)
    print(f"{plik.name}: skopiowano z {SZABLON_CRAZYGAMES_ZRODLO.name}")


ARKUSZ_CENNIK = "Traffic cennik"


def _waluta_z_formatu(number_format: str) -> str:
    """Kolumna Stawka nie ma osobnej kolumny Waluta - walutę koduje samo
    formatowanie liczby (np. "$"#,##0.00 vs €#,##0.00). Domyślnie USD, jeśli
    formatowanie nie wskazuje inaczej."""
    if "€" in number_format or "EUR" in number_format.upper():
        return "EUR"
    if "zł" in number_format or "PLN" in number_format.upper():
        return "PLN"
    return "USD"


def export_cennik_wydawcow(numery_xlsx=None):
    """Cennik placementów wydawców zewnętrznych (KIDOZ/PRADO/Adverty/Odeeo/
    Crazygames/POKI/...), z zakładki "Traffic cennik" w Numery_zlecen_2026.xlsx
    (nie per-akant - cennik jest wspólny). Kolumny: A=Wydawca, B=Format,
    C=Stawka (waluta z formatowania komórki, patrz _waluta_z_formatu),
    D=Uwagi (informacyjne, nieużywane przez generator). Wiersz 1=nagłówki,
    dane od wiersza 2 - bez banera/kolumny Aktywny, w odróżnieniu od innych
    zakładek tego pliku.

    Brak zakładki (jeszcze nie dodana / stary plik) daje pusty cennik, nie
    błąd - żeby nie blokować startu aplikacji, zanim dział traffic ją
    uzupełni."""
    wb = openpyxl.load_workbook(numery_xlsx or NUMERY_XLSX, data_only=True)

    wynik = {}
    if ARKUSZ_CENNIK in wb.sheetnames:
        ws = wb[ARKUSZ_CENNIK]
        r = 2
        while True:
            wydawca = ws.cell(row=r, column=1).value
            format_placementu = ws.cell(row=r, column=2).value
            if not (wydawca and str(wydawca).strip()) and not (format_placementu and str(format_placementu).strip()):
                break
            if wydawca and format_placementu:
                komorka_stawki = ws.cell(row=r, column=3)
                cena = komorka_stawki.value
                if cena is not None:
                    wynik.setdefault(str(wydawca).strip(), {})[str(format_placementu).strip()] = {
                        "cena": float(cena),
                        "waluta": _waluta_z_formatu(komorka_stawki.number_format),
                    }
            r += 1

    with open(DATA_OUT / "cennik_wydawcow.json", "w", encoding="utf-8") as f:
        json.dump(wynik, f, ensure_ascii=False, indent=2)
    print(f"cennik_wydawcow.json: {sum(len(v) for v in wynik.values())} aktywnych stawek u {len(wynik)} wydawców")


def _column_values(ws, col_letter, start_row):
    values = []
    r = start_row
    while True:
        v = ws[f"{col_letter}{r}"].value
        if v is None or (isinstance(v, str) and not v.strip()):
            break
        values.append(str(v).strip())
        r += 1
    return values


MODELE_SPRZEDAZY_WALIDNE = ["CPM", "CPC", "CPV", "FF"]


def export_slowniki():
    """Listy rozwijane z arkusza 'Dane' pliku kampanii Igora.

    Kolumna F arkusza 'Dane' ma więcej pozycji ("Liczba wyświetleń" itd.) niż
    faktycznie obowiązujących modeli sprzedaży — potwierdzone, że jedyne
    prawdziwe modele to CPM/CPC/CPV/FF, więc listę obcinamy do nich.

    Lista klientów NIE jest tu już eksportowana jako płaska globalna lista —
    zastąpiona przez klienci_agencyjni.json (scoped per account+agencja/
    klient bezpośredni), patrz export_klienci_agencyjni().
    """
    wb = openpyxl.load_workbook(KAMPANIE_XLSM, data_only=True)
    ws = wb["Dane"]

    wszystkie_modele = _column_values(ws, "F", 2)
    modele = [m for m in MODELE_SPRZEDAZY_WALIDNE if m in wszystkie_modele]

    slowniki = {
        "przejsciowa": _column_values(ws, "B", 2),
        "format_reklamowy": _column_values(ws, "C", 2),
        "target": _column_values(ws, "D", 2),
        "podmiot": _column_values(ws, "E", 2),
        "model_sprzedazy": modele,
    }

    with open(DATA_OUT / "slowniki.json", "w", encoding="utf-8") as f:
        json.dump(slowniki, f, ensure_ascii=False, indent=2)
    print(f"slowniki.json: {[(k, len(v)) for k, v in slowniki.items()]}")


if __name__ == "__main__":
    DATA_OUT.mkdir(parents=True, exist_ok=True)
    export_podmioty()
    export_klienci_agencyjni()
    export_terminy_platnosci_klientow()
    export_mediafarm()
    export_slowniki()
    export_szablony_wydawcow()
    export_szablon_crazygames()
    export_cennik_wydawcow()
