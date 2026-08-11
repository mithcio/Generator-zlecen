"""Generuje IO/brief dla wydawcy zewnętrznego (KIDOZ / Adverty / Odeeo /
Crazygames / POKI) — dla KIDOZ/Adverty/Odeeo/Crazygames kopiuje gotowy
szablon (app/data/szablony_wydawcow/*.xlsx) i podmienia w nim tylko komórki,
które w oryginale odwoływały się do innych zakładek (Kampania!/Zlecenie!/
Podmioty!) albo do danych, które teraz mamy we własnym systemie — reszta
formatowania i formuł wewnętrznych (odwołujących się tylko do komórek w tym
samym arkuszu) zostaje nietknięta. POKI nie ma oryginalnego szablonu do
skopiowania (submituje się mailem, nie plikiem) — budowany od zera jako
prosta lista pole/wartość.

Ceny/CPM pochodzą z app/services/cennik.py (zakładka „Traffic cennik” w
Numery_zlecen_2026.xlsx) - jeśli dla (wydawca, format) nie ma tam aktywnej
stawki, generowanie się nie uda (BladCennika) zamiast po cichu wpisać 0.

KIDOZ ma w cenniku "bliźniaczy" wpis pod wydawcą "PRADO" - ten sam plik/IO,
inna stawka dla targetu Adults/Silver (patrz _kidoz_wydawca_cennika).
"""
from datetime import date
from pathlib import Path
from shutil import copyfile

import openpyxl
from openpyxl.styles import Font

from app.models.podmiot import SpolkaMediafarm
from app.models.zlecenie import Zlecenie
from app.services import cennik
from app.services.eksport_nazwy import oczysc_nazwe
from app.services.lookup_podmiotu import formatuj_telefon

SZABLONY_DIR = Path(__file__).resolve().parent.parent / "data" / "szablony_wydawcow"

WYDAWCY_OBSLUGIWANI = ["KIDOZ", "Adverty", "Odeeo", "Crazygames", "POKI"]

# Placementy POKI (dropdown w kroku 5, widoczny tylko gdy wybrano POKI) - w
# cenniku to osobne wiersze pod wydawcą "POKI", zwykłe nazwy formatu (bez
# prefiksu) - tak jak faktycznie wygląda zakładka „Traffic cennik”.
PLACEMENTY_POKI = ["ImViTa", "Overlay", "HPTO", "Rewarded"]

# Które kategorie formatu (Format reklamowy z kroku 2) pasują do których
# wydawców zewnętrznych - ustalone z użytkownikiem, do ostrzegania przy
# generowaniu (patrz czy_format_pasuje), nie do blokowania. PRADO nie
# występuje tu osobno - to ten sam wybór "KIDOZ" w dropdownie, tylko inna
# stawka w cenniku zależnie od targetu (patrz _kidoz_wydawca_cennika).
WYDAWCY_DLA_KATEGORII = {
    "rewarded": {"KIDOZ", "POKI", "CRAZYGAMES"},
    "dooh": {"ADVERTY"},
    "audio": {"ODEEO"},
    "interstitial": {"KIDOZ"},
    "non_standard": {"POKI"},
}


def _kategoria_formatu(format_reklamowy: str) -> str | None:
    """None = format spoza wszystkich powyższych kategorii (YouTube, Display
    CPC, Mailing, Roblox DISPLAY/VIDEO, Video interaktywne...) - nie pasuje
    do żadnego obsługiwanego wydawcy zewnętrznego."""
    if format_reklamowy.startswith("Rewarded"):
        return "rewarded"
    if format_reklamowy.startswith("In-game DOOH"):
        return "dooh"
    if format_reklamowy.startswith("In-game audio"):
        return "audio"
    if format_reklamowy.startswith(("Interstitial", "Mobistitial")):
        return "interstitial"
    if format_reklamowy == "Non-standard":
        return "non_standard"
    return None


def czy_format_pasuje(wydawca_zewnetrzny: str, format_reklamowy: str) -> bool:
    """Czy Format reklamowy z kroku 2 zwyczajowo pasuje do wybranego
    wydawcy zewnętrznego - do ostrzeżenia w kroku 5 przy próbie
    wygenerowania pliku, nie do twardej blokady."""
    kategoria = _kategoria_formatu(format_reklamowy)
    if kategoria is None:
        return False
    return wydawca_zewnetrzny in WYDAWCY_DLA_KATEGORII[kategoria]


class BladSzablonuWydawcy(Exception):
    """Brak pliku szablonu (przenieś się/uruchom ponownie
    scripts/export_seed_data.py) — do poprawienia przez użytkownika, nie
    błąd programu."""


def _skopiuj_szablon(nazwa_pliku_szablonu: str, sciezka_docelowa: Path) -> openpyxl.Workbook:
    szablon = SZABLONY_DIR / f"{nazwa_pliku_szablonu}.xlsx"
    if not szablon.exists():
        raise BladSzablonuWydawcy(
            f"Brak szablonu {szablon}. Uruchom scripts/export_seed_data.py, "
            "żeby go wygenerować/skopiować ze źródeł."
        )
    sciezka_docelowa.parent.mkdir(parents=True, exist_ok=True)
    copyfile(szablon, sciezka_docelowa)
    return openpyxl.load_workbook(sciezka_docelowa)


def _czy_display(format_reklamowy: str) -> bool:
    return "DISPLAY" in format_reklamowy.upper()


def _czy_video(format_reklamowy: str) -> bool:
    return "VIDEO" in format_reklamowy.upper()


def _adverty_placement(format_reklamowy: str) -> str:
    """Klucz w cenniku (wydawca "Adverty") - "Display" albo "Video", nie
    dokładny format_reklamowy (cennik nie rozróżnia KIDS/ADULTS)."""
    return "Display" if _czy_display(format_reklamowy) else "Video"


def _target_z_formatu(format_reklamowy: str) -> str:
    """"KIDS" albo "ADULTS", odczytane z końcówki nazwy formatu (np.
    "Interstitial KIDS" -> "KIDS") - jedyne, co ma trafiać w pole Targeting
    w IO wydawcy. Uwagi (Zlecenie.pola.uwagi) to nasze wewnętrzne informacje
    do dokumentu dla klienta i NIE MOGĄ trafiać do żadnego pliku dla
    wydawcy zewnętrznego - patrz ustalenia z użytkownikiem."""
    gorny = format_reklamowy.upper()
    if gorny.endswith("KIDS"):
        return "KIDS"
    if gorny.endswith("ADULTS"):
        return "ADULTS"
    return ""


def _kidoz_typ_kampanii(format_reklamowy: str) -> str:
    """"Interstitial" dla formatów displayowych (Interstitial/Mobistitial i
    wariacje), "Rewarded" dla wszystkiego innego (rewarded, in-app video,
    pre-roll...) - patrz ustalenia z użytkownikiem, szersze niż oryginalna
    formuła szablonu (ta nie uwzględniała w ogóle Mobistitial). To też klucz
    placementu w cenniku (kolumna Format, u KIDOZ/PRADO)."""
    if format_reklamowy.startswith(("Interstitial", "Mobistitial")):
        return "Interstitial"
    return "Rewarded"


def _kidoz_wydawca_cennika(target: str) -> str:
    """Ten sam plik/Insertion Order co KIDOZ, ale dla targetu Adults/Silver
    cena jest inna - w cenniku figuruje pod osobnym wydawcą "PRADO" (patrz
    ustalenia z użytkownikiem). Sam wygenerowany plik/nazwa zostają jak dla
    KIDOZ - zmienia się tylko to, skąd bierzemy stawkę."""
    return "PRADO" if target.upper() in ("ADULTS", "SILVER") else "KIDOZ"


def _kidoz_agencja_wlasciciel(podmiot_realizujacy: str) -> str:
    if podmiot_realizujacy == "Sp. k.":
        return "Mediafarm"
    if podmiot_realizujacy == "Sp. z o.o.":
        return "Mediafarm Spółka"
    return ""


def _capp_tekst(capping: int | None) -> str:
    return f"Capp {capping}" if capping is not None else "Capp brak"


def _fmt_data(d: date) -> str:
    return d.strftime("%d/%m/%Y")


# --- nazwy plików ---------------------------------------------------------
# KIDOZ/Adverty/Odeeo/Crazygames mają, ustaloną z użytkownikiem, konkretną
# pisownię nazwy wydawcy w nazwie pliku - inną niż klucz w GENERATORY (który
# odpowiada wartościom z dropdowna "Wydawcy zewnętrzni", zawsze wielkimi
# literami).
NAZWA_W_PLIKU = {"KIDOZ": "KIDOZ", "ADVERTY": "Adverty", "ODEEO": "Odeeo", "CRAZYGAMES": "Crazygames"}


def _nazwa_pliku_purchase(wydawca: str, nr_zlecenia: str, format_reklamowy: str, nazwa_kampanii: str) -> str:
    # np. Adverty_Purchase_K-2026-020_In-game_DOOH_DISPLAY_ADULTS_Starcom_LEGO...
    return oczysc_nazwe(
        f"{NAZWA_W_PLIKU[wydawca]}_Purchase_{nr_zlecenia}_{format_reklamowy}_{nazwa_kampanii}"
    )


def _nazwa_pliku_crazygames(klient: str, nazwa_kampanii: str) -> str:
    return oczysc_nazwe(f"Crazygames_Mediafarm_{klient}_{nazwa_kampanii}_IO")


def _nazwa_pliku_poki(nazwa_kampanii: str) -> str:
    return oczysc_nazwe(f"POKI_{nazwa_kampanii}_IO")


# --- KIDOZ / Adverty / Odeeo -----------------------------------------------

def generuj_kidoz(zlecenie: Zlecenie, spolka: SpolkaMediafarm, kontakt_accounta: dict, folder: Path) -> Path:
    pola = zlecenie.pola
    stawka = cennik.stawka(_kidoz_wydawca_cennika(pola.target), _kidoz_typ_kampanii(pola.format_reklamowy))
    nazwa = _nazwa_pliku_purchase("KIDOZ", pola.nr_zlecenia, pola.format_reklamowy, pola.nazwa_kampanii)
    sciezka = folder / f"{nazwa}.xlsx"

    wb = _skopiuj_szablon("kidoz", sciezka)
    ws = wb.active

    ws["D5"] = _kidoz_agencja_wlasciciel(pola.podmiot_realizujacy)
    ws["D6"] = pola.account_manager
    ws["D7"] = kontakt_accounta.get("email", "")
    ws["D8"] = pola.klient
    ws["D9"] = stawka.waluta
    ws["D10"] = pola.nr_zlecenia

    ws["C14"] = pola.nazwa_kampanii
    ws["D14"] = zlecenie.data_startu
    ws["E14"] = zlecenie.data_konca
    ws["G14"] = _kidoz_typ_kampanii(pola.format_reklamowy)
    ws["I14"] = round(zlecenie.liczba_total)
    ws["J14"] = stawka.cena
    ws["N14"] = _target_z_formatu(pola.format_reklamowy)  # Campaign Targeting - NIE Uwagi
    ws["O14"] = _capp_tekst(pola.capping)

    wb.save(sciezka)
    return sciezka


def generuj_adverty(zlecenie: Zlecenie, spolka: SpolkaMediafarm, kontakt_accounta: dict, folder: Path) -> Path:
    pola = zlecenie.pola
    stawka = cennik.stawka("Adverty", _adverty_placement(pola.format_reklamowy))
    nazwa = _nazwa_pliku_purchase("ADVERTY", pola.nr_zlecenia, pola.format_reklamowy, pola.nazwa_kampanii)
    sciezka = folder / f"{nazwa}.xlsx"

    wb = _skopiuj_szablon("adverty", sciezka)
    ws = wb.active

    ws["B8"] = spolka.nazwa
    ws["B9"] = pola.dom_mediowy
    ws["B10"] = pola.klient
    ws["B11"] = pola.nazwa_kampanii
    ws["B12"] = zlecenie.data_startu
    ws["B13"] = zlecenie.data_konca
    ws["B14"] = round(zlecenie.liczba_total)
    ws["B15"] = stawka.cena

    # Display/Video ustalane wprost z formatu (nie z ceny, jak w oryginale -
    # przy cenie=0 stara formuła zawsze dałaby "N"/FALSE dla obu).
    ws["B19"] = "Y" if _czy_display(pola.format_reklamowy) else "N"
    ws["B20"] = "Y" if _czy_video(pola.format_reklamowy) else "N"
    # C19/C20 (impression breakdown) zostają formułami w arkuszu (=IF(B19=...))
    # - przeliczą się same na 100%/0% zależnie od B19/B20.

    ws["B30"] = _target_z_formatu(pola.format_reklamowy)  # Targeting Details - NIE Uwagi

    wb.save(sciezka)
    return sciezka


def _wysokosc_wiersza_dla_tekstu(tekst: str, szerokosc_kolumny: float, wysokosc_min: float = 30.6) -> float:
    """A14 (Odeeo) ma włączone zawijanie tekstu (wrap_text), ale pełna nazwa
    kampanii (nazwa_kampanii_nr_zlecenia) bywa dłuższa niż mieści się w
    domyślnej wysokości wiersza - dobieramy wysokość do liczby linii, na
    które tekst się zawinie."""
    znakow_w_linii = max(int(szerokosc_kolumny), 1)
    liczba_linii = max(1, -(-len(tekst) // znakow_w_linii))  # ceil
    return max(wysokosc_min, liczba_linii * 15.0)


def generuj_odeeo(zlecenie: Zlecenie, spolka: SpolkaMediafarm, kontakt_accounta: dict, folder: Path) -> Path:
    pola = zlecenie.pola
    stawka = cennik.stawka("Odeeo", "In-Game Audio")
    nazwa = _nazwa_pliku_purchase("ODEEO", pola.nr_zlecenia, pola.format_reklamowy, pola.nazwa_kampanii)
    sciezka = folder / f"{nazwa}.xlsx"

    wb = _skopiuj_szablon("odeeo", sciezka)
    ws = wb.active

    ws["D3"] = spolka.nazwa
    ws["D4"] = spolka.adres
    ws["D5"] = formatuj_telefon(kontakt_accounta.get("telefon"))
    ws["D6"] = kontakt_accounta.get("email", "")
    ws["D7"] = pola.account_manager

    # Kontakt po stronie Odeeo (B6 email, B7 nazwisko) - zostawiamy puste,
    # uzupełniane ręcznie później.
    ws["B6"] = None
    ws["B7"] = None

    nazwa_pelna = f"{pola.nazwa_kampanii}_{pola.nr_zlecenia}"
    ws["A14"] = nazwa_pelna
    ws["B14"] = f"{_fmt_data(zlecenie.data_startu)} - {_fmt_data(zlecenie.data_konca)}"
    ws["C14"] = round(zlecenie.liczba_total)
    ws["D14"] = stawka.cena
    # E14 (Total Net Cost) zostaje formułą w arkuszu (=D14*C14/1000).
    ws.row_dimensions[14].height = _wysokosc_wiersza_dla_tekstu(
        nazwa_pelna, ws.column_dimensions["A"].width
    )

    ws["E17"] = pola.account_manager
    # E16 zostaje formułą w arkuszu (=D3).
    ws["E19"] = date.today()  # data wystawienia dokumentu

    wb.save(sciezka)
    return sciezka


# --- Crazygames -------------------------------------------------------------

def generuj_crazygames(zlecenie: Zlecenie, spolka: SpolkaMediafarm, kontakt_accounta: dict, folder: Path) -> Path:
    """Szablon jest już gotowym, prostym plikiem (nie wycinkiem z dużego
    xlsm) - ustalone z użytkownikiem, że poza kilkoma polami formatowanie i
    reszta wartości (Campaign Type, Format="Video Pre-Roll", Size,
    Targeting Options, formuła Budget) zostają bez zmian. Placement jest
    zawsze jeden ("Video Pre-Roll") - w cenniku figuruje pod wydawcą
    "Crazygames", placement "Video"."""
    pola = zlecenie.pola
    stawka = cennik.stawka("Crazygames", "Video")
    nazwa = _nazwa_pliku_crazygames(pola.klient, pola.nazwa_kampanii)
    sciezka = folder / f"{nazwa}.xlsx"

    wb = _skopiuj_szablon("crazygames", sciezka)
    ws = wb.active

    ws["B4"] = pola.klient  # Advertiser
    ws["C4"] = zlecenie.data_startu  # Start Date
    ws["D4"] = zlecenie.data_konca  # End Date
    ws["G4"] = round(zlecenie.liczba_total)  # Impressions
    ws["I4"] = stawka.cena  # CPM (komórka ma już format "[$€]#,##0.00")
    # J4 (Budget) zostaje formułą w arkuszu (=G4/1000*I4).
    capping_tekst = f"{pola.capping}/campaign" if pola.capping is not None else "brak/campaign"
    ws["K4"] = capping_tekst

    wb.save(sciezka)
    return sciezka


# --- POKI --------------------------------------------------------------

POLA_POKI = ["Campaign", "Timing", "Format", "Impressions", "Sites", "Target", "Geo", "Capp", "Device", "CPM"]


def generuj_poki(
    zlecenie: Zlecenie, spolka: SpolkaMediafarm, kontakt_accounta: dict, folder: Path, placement: str
) -> Path:
    """POKI nie dostaje pliku wg gotowego szablonu (brief wysyłany mailem) -
    budujemy od zera prostą listę pole/wartość: kolumna A pogrubione
    etykiety (Campaign..CPM), kolumna B wartości. "Format" to nowy wiersz
    (nazwa placementu POKI po angielsku), w cenniku szukany pod wydawcą
    "POKI"."""
    pola = zlecenie.pola
    stawka = cennik.stawka("POKI", placement)
    nazwa = _nazwa_pliku_poki(pola.nazwa_kampanii)
    sciezka = folder / f"{nazwa}.xlsx"

    impresje = f"{round(zlecenie.liczba_total):,}".replace(",", " ") + " PVs"
    timing = f"{_fmt_data_kropki(zlecenie.data_startu)}-{_fmt_data_kropki(zlecenie.data_konca)}"
    capp_tekst = str(pola.capping) if pola.capping is not None else "brak"

    wartosci = {
        "Campaign": pola.nazwa_kampanii,
        "Timing": timing,
        "Format": placement,
        "Impressions": impresje,
        "Sites": "poki.pl",
        "Target": pola.target.capitalize(),
        "Geo": "Poland",
        "Capp": capp_tekst,
        "Device": "cross-device",
        "CPM": stawka.sformatowana(),
    }

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "POKI"

    max_a, max_b = 0, 0
    for i, etykieta in enumerate(POLA_POKI, start=1):
        c_etykieta = ws.cell(row=i, column=1, value=etykieta)
        c_etykieta.font = Font(bold=True)
        wartosc = wartosci[etykieta]
        ws.cell(row=i, column=2, value=wartosc)
        max_a = max(max_a, len(etykieta))
        max_b = max(max_b, len(str(wartosc)))
    ws.column_dimensions["A"].width = max_a + 2
    ws.column_dimensions["B"].width = max_b + 2

    sciezka.parent.mkdir(parents=True, exist_ok=True)
    wb.save(sciezka)
    return sciezka


def _fmt_data_kropki(d: date) -> str:
    return d.strftime("%d.%m.%Y")


GENERATORY = {
    "KIDOZ": generuj_kidoz,
    "ADVERTY": generuj_adverty,
    "ODEEO": generuj_odeeo,
    "CRAZYGAMES": generuj_crazygames,
}
