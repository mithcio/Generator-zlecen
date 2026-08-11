"""Generuje plik xlsx Zlecenia na podstawie wspólnego układu z
app/templates/zlecenie_layout.py.

Większość pól to gotowe wartości (jak ZamienFormuly w starym VBA — bo
odwołują się do danych spoza tego arkusza: Podmioty, Kampania z innego
pliku). Wyjątek: liczba wyświetleń/klików/odtworzeń (4.5) i kwota brutto
(7.3) to prawdziwe formuły Excela odwołujące się do koszt jednostkowy (4.6)
i koszt netto (7.1) — dzięki temu account może w Excelu poprawić np. budżet
netto i od razu zobaczyć przeliczoną liczbę i kwotę brutto.

Wygląd (kolory, szerokości kolumn, ramki, logo, jedna strona) odwzorowuje
oryginalny szablon Zlecenie z Generator_zleceń_i_danych.xlsm.
"""
from pathlib import Path

from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.worksheet.worksheet import Worksheet

from app.models.podmiot import DanePodmiotu, SpolkaMediafarm
from app.models.zlecenie import Zlecenie
from app.templates.zlecenie_layout import (
    LiniaPodpisu,
    Naglowek,
    NaglowekZIdentyfikatorem,
    Pozycja,
    Tekst,
    zbuduj_layout,
)

LOGO_PATH = Path(__file__).resolve().parent.parent / "assets" / "logo_mediafarm.png"

KOLOR_NAGLOWEK = "000080"
KOLOR_ETYKIETA = "D9D9D9"

FILL_NAGLOWEK = PatternFill("solid", fgColor=KOLOR_NAGLOWEK)
FILL_ETYKIETA = PatternFill("solid", fgColor=KOLOR_ETYKIETA)

FONT_NAGLOWEK = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
FONT_ETYKIETA = Font(name="Calibri", size=11, bold=True, color="000000")
FONT_WARTOSC = Font(name="Calibri", size=11, bold=False, color="000000")
FONT_TEKST = Font(name="Calibri", size=10, bold=False, color="000000")

CIENKI = Side(style="thin", color="000000")
GRUBY = Side(style="medium", color="000000")

ROW_HEIGHT = 22.5


def _ramka(cell, left=None, top=None, right=None, bottom=None) -> None:
    """Ustawia tylko wskazane boki ramki, zachowując pozostałe."""
    b = cell.border
    cell.border = Border(
        left=left or b.left, top=top or b.top, right=right or b.right, bottom=bottom or b.bottom,
    )


def _wstaw_logo(ws: Worksheet) -> None:
    if not LOGO_PATH.exists():
        return
    img = XLImage(str(LOGO_PATH))
    img.width, img.height = 112, 90
    ws.add_image(img, "A1")
    ws.row_dimensions[1].height = 68


def _wiersz_naglowek_z_id(ws: Worksheet, r: int, element: NaglowekZIdentyfikatorem) -> None:
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=2)
    ws.merge_cells(start_row=r, start_column=3, end_row=r, end_column=4)
    c_tekst = ws.cell(row=r, column=1, value=element.tekst)
    c_etykieta = ws.cell(row=r, column=3, value=element.etykieta_id)
    c_wartosc = ws.cell(row=r, column=5, value=element.wartosc_id)
    for c in (c_tekst, c_etykieta):
        c.fill = FILL_NAGLOWEK
        c.font = FONT_NAGLOWEK
        c.alignment = Alignment(vertical="center", wrap_text=True)
    c_wartosc.font = Font(name="Calibri", size=11, bold=True, color="000000")
    c_wartosc.alignment = Alignment(vertical="center")
    for col in range(1, 6):
        cell = ws.cell(row=r, column=col)
        _ramka(cell, left=CIENKI, top=CIENKI, right=CIENKI, bottom=CIENKI)
    ws.row_dimensions[r].height = ROW_HEIGHT


def _wiersz_naglowek(ws: Worksheet, r: int, element: Naglowek) -> None:
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=5)
    cell = ws.cell(row=r, column=1, value=element.tekst)
    cell.fill = FILL_NAGLOWEK
    cell.font = FONT_NAGLOWEK
    cell.alignment = Alignment(vertical="center", wrap_text=True)
    for col in range(1, 6):
        _ramka(ws.cell(row=r, column=col), left=CIENKI, top=CIENKI, right=CIENKI, bottom=CIENKI)
    ws.row_dimensions[r].height = ROW_HEIGHT


def _wiersz_pozycja(ws: Worksheet, r: int, element: Pozycja) -> None:
    ws.merge_cells(start_row=r, start_column=3, end_row=r, end_column=5)
    c_numer = ws.cell(row=r, column=1, value=element.numer)
    c_etykieta = ws.cell(row=r, column=2, value=element.etykieta)
    c_wartosc = ws.cell(row=r, column=3, value=element.wartosc)
    for c in (c_numer, c_etykieta):
        c.fill = FILL_ETYKIETA
        c.font = FONT_ETYKIETA
        c.alignment = Alignment(vertical="center", wrap_text=True)
    c_wartosc.font = FONT_WARTOSC
    # Wyrównanie do lewej jawnie - inaczej Excel domyślnie wyrównuje liczby
    # (4.5/4.6/7.1-7.3, prawdziwe wartości liczbowe) do prawej, a resztę
    # kolumny C (tekst) do lewej, co wygląda na rozjazd w jednej kolumnie.
    c_wartosc.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    for col in range(1, 6):
        _ramka(ws.cell(row=r, column=col), left=CIENKI, top=CIENKI, right=CIENKI, bottom=CIENKI)
    ws.row_dimensions[r].height = ROW_HEIGHT


def _wiersz_tekst(ws: Worksheet, r: int, element: Tekst) -> None:
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=5)
    cell = ws.cell(row=r, column=1, value=element.tresc)
    cell.font = FONT_TEKST
    cell.alignment = Alignment(vertical="top", wrap_text=True)
    for col in range(1, 6):
        _ramka(ws.cell(row=r, column=col), left=CIENKI, top=CIENKI, right=CIENKI, bottom=CIENKI)
    # Przybliżona wysokość dla dłuższych akapitów (np. 8.2), żeby tekst się mieścił.
    znaki_w_wierszu = 118
    liczba_linii = max(1, -(-len(element.tresc) // znaki_w_wierszu))
    ws.row_dimensions[r].height = max(ROW_HEIGHT, liczba_linii * 13)


def _wiersz_podpis(ws: Worksheet, r: int, element: LiniaPodpisu) -> int:
    """Wiersz z etykietami "Podpis i pieczęć..." + osobny, wysoki, w pełni
    oprawiony w ramkę wiersz PONIŻEJ na samo miejsce na podpis/pieczątkę - tak
    jak w oryginalnym szablonie (pusty oprawiony prostokąt POD etykietami, nie
    nad nimi - i musi mieć widoczną ramkę, nie tylko zarezerwowaną wysokość)."""
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=2)
    ws.merge_cells(start_row=r, start_column=3, end_row=r, end_column=5)
    c_lewa = ws.cell(row=r, column=1, value=element.lewa)
    c_prawa = ws.cell(row=r, column=3, value=element.prawa)
    for c in (c_lewa, c_prawa):
        c.font = FONT_ETYKIETA
        c.alignment = Alignment(horizontal="center", vertical="center")
    for col in range(1, 6):
        _ramka(ws.cell(row=r, column=col), left=CIENKI, top=CIENKI, right=CIENKI, bottom=CIENKI)
    ws.row_dimensions[r].height = ROW_HEIGHT

    r_pieczec = r + 1
    ws.merge_cells(start_row=r_pieczec, start_column=1, end_row=r_pieczec, end_column=5)
    for col in range(1, 6):
        _ramka(ws.cell(row=r_pieczec, column=col), left=CIENKI, top=CIENKI, right=CIENKI, bottom=CIENKI)
    ws.row_dimensions[r_pieczec].height = 55  # miejsce na pieczątkę/podpis
    return r_pieczec


FORMAT_KWOTY = '#,##0.00" PLN"'
FORMAT_LICZBY = "#,##0"
FORMAT_PROCENT = "0%"


def _wstaw_formuly(ws: Worksheet, wiersze: dict[str, int], zlecenie: Zlecenie) -> None:
    """4.5 (liczba), 4.6 (koszt jednostkowy), 7.1 (netto), 7.2 (VAT) i 7.3
    (brutto) muszą być prawdziwymi liczbami/formułami Excela — nie
    pre-sformatowanym tekstem — żeby account mógł w Excelu poprawić np.
    budżet netto i mieć automatyczne przeliczenie liczby i kwoty brutto,
    zamiast ręcznie liczyć wszystko od nowa."""
    pola = zlecenie.pola
    r46, r45 = wiersze.get("4.6"), wiersze.get("4.5")
    r71, r72, r73 = wiersze.get("7.1"), wiersze.get("7.2"), wiersze.get("7.3")

    def _ustaw(wiersz: int | None, wartosc, format_liczby: str) -> None:
        if wiersz is None:
            return
        c = ws.cell(row=wiersz, column=3, value=wartosc)
        c.number_format = format_liczby
        c.font = FONT_WARTOSC
        c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

    r43 = wiersze.get("4.3")
    if r43 is not None and pola.capping is not None:
        # Capping jako prawdziwa liczba (nie tekst z ogólnego layoutu Pozycja),
        # żeby dało się w przyszłości użyć jej w formule Excela. "brak" (None)
        # zostaje tekstem - nie ma sensownej wartości liczbowej.
        _ustaw(r43, pola.capping, FORMAT_LICZBY)

    _ustaw(r71, round(zlecenie.budzet_total, 2), FORMAT_KWOTY)
    _ustaw(r72, 0.23, FORMAT_PROCENT)
    if r73 and r71 and r72:
        _ustaw(r73, f"=C{r71}*(1+C{r72})", FORMAT_KWOTY)

    # FF (opłata stała) nie ma kosztu jednostkowego — cały budżet to jedna
    # opłata, więc 4.6 pokazuje wprost budżet, a 4.5 (liczba) zawsze = 1.
    if pola.model_sprzedazy == "FF":
        _ustaw(r46, round(zlecenie.budzet_total, 2), FORMAT_KWOTY)
        _ustaw(r45, 1, FORMAT_LICZBY)
    else:
        _ustaw(r46, round(pola.koszt_jednostkowy, 2), FORMAT_KWOTY)
        if r45 and r46 and r71:
            mnoznik = "*1000" if pola.model_sprzedazy == "CPM" else ""
            _ustaw(r45, f"=IF(C{r46}=0,0,C{r71}/C{r46}{mnoznik})", FORMAT_LICZBY)


def _pogrub_ramke_zewnetrzna(ws: Worksheet, pierwszy_wiersz: int, ostatni_wiersz: int) -> None:
    for r in range(pierwszy_wiersz, ostatni_wiersz + 1):
        _ramka(ws.cell(row=r, column=1), left=GRUBY)
        _ramka(ws.cell(row=r, column=5), right=GRUBY)
    for col in range(1, 6):
        _ramka(ws.cell(row=pierwszy_wiersz, column=col), top=GRUBY)
        _ramka(ws.cell(row=ostatni_wiersz, column=col), bottom=GRUBY)


def generuj_xlsx(
    zlecenie: Zlecenie,
    podmiot: DanePodmiotu,
    spolka: SpolkaMediafarm,
    kontakt_accounta: dict,
    sciezka: Path,
) -> Path:
    layout = zbuduj_layout(zlecenie, podmiot, spolka, kontakt_accounta)

    wb = Workbook()
    ws = wb.active
    ws.title = "Zlecenie"
    ws.sheet_view.showGridLines = False

    ws.column_dimensions["A"].width = 5
    ws.column_dimensions["B"].width = 35
    ws.column_dimensions["C"].width = 6.13
    ws.column_dimensions["D"].width = 15.4
    ws.column_dimensions["E"].width = 65

    _wstaw_logo(ws)

    r = 2
    pierwszy_wiersz_tabeli = r
    wiersz_po_numerze: dict[str, int] = {}
    for element in layout:
        if isinstance(element, NaglowekZIdentyfikatorem):
            _wiersz_naglowek_z_id(ws, r, element)
            r += 1
        elif isinstance(element, Naglowek):
            _wiersz_naglowek(ws, r, element)
            r += 1
        elif isinstance(element, Pozycja):
            _wiersz_pozycja(ws, r, element)
            wiersz_po_numerze[element.numer] = r
            r += 1
        elif isinstance(element, Tekst):
            _wiersz_tekst(ws, r, element)
            r += 1
        elif isinstance(element, LiniaPodpisu):
            r = _wiersz_podpis(ws, r, element)
            r += 1

    _wstaw_formuly(ws, wiersz_po_numerze, zlecenie)
    _pogrub_ramke_zewnetrzna(ws, pierwszy_wiersz_tabeli, r - 1)

    ws.page_setup.orientation = "portrait"
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 1
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.print_area = f"A1:E{r - 1}"

    sciezka.parent.mkdir(parents=True, exist_ok=True)
    wb.save(sciezka)
    return sciezka
