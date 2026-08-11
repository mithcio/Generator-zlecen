"""Rezerwacja numeru zlecenia bezpośrednio w pliku Numery_zlecen_2026.xlsx
(ścieżka ustawiana w panelu Ustawienia — ikona koła zębatego).

Plik jest współdzielony (żyje w folderze zsynchronizowanym z chmurą, żeby
kilku akantów mogło z niego korzystać naraz), więc żeby zminimalizować okno
na konflikt, rezerwacja to jedna szybka sekwencja: otwórz -> znajdź pierwszy
wolny numer -> zapisz status/datę/akanta -> zapisz i zamknij od razu. Numer
trafia do pamięci programu i od tej chwili plik jest znów wolny dla innych.
To nie jest twarda blokada (nie ma serwera plików pilnującego dostępu) -
przy dosłownie jednoczesnym kliknięciu dwóch osób nadal jest teoretyczne
ryzyko wyścigu, ale krótkie okno otwarcia/zapisu/zamknięcia drastycznie je
ogranicza względem trzymania pliku otwartego przez cały czas pracy w
kreatorze.

Arkusz "Numery zleceń": kolumny A-D = SP. K. (numer/status/data/użytkownik),
kolumny F-I = SP. Z O.O. (to samo przesunięte o kolumnę E jako separator).
"""
from datetime import date
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment
from openpyxl.worksheet.worksheet import Worksheet

from app.services import ustawienia

ARKUSZ = "Numery zleceń"

PREFIKSY = {"Sp. k.": "K", "Sp. z o.o.": "S"}

KOLUMNY = {
    "Sp. k.": {"numer": "A", "status": "B", "data": "C", "akant": "D"},
    "Sp. z o.o.": {"numer": "F", "status": "G", "data": "H", "akant": "I"},
}

WYROWNANIE_SRODEK = Alignment(horizontal="center")


class BladNumeracji(Exception):
    """Ścieżka do pliku nieustawiona/nieprawidłowa, brak zakładki, brak
    wolnych numerów, plik zajęty przez inny program — wszystko, co
    użytkownik musi sam poprawić (w Ustawieniach albo w samym pliku), a nie
    błąd programu."""


def _sprawdz_podmiot(podmiot_realizujacy: str) -> None:
    if podmiot_realizujacy not in KOLUMNY:
        raise ValueError(
            f"Nieznany podmiot realizujący: {podmiot_realizujacy!r} "
            f"(oczekiwano: {', '.join(KOLUMNY)})"
        )


def _otworz_workbook(sciezka: str) -> openpyxl.Workbook:
    try:
        return openpyxl.load_workbook(sciezka)
    except OSError as err:
        # Najczęstsza przyczyna: plik otwarty w tej chwili w Excelu (albo
        # jeszcze w trakcie synchronizacji z chmurą) - blokuje odczyt/zapis.
        raise BladNumeracji(
            f"Nie udało się otworzyć pliku {sciezka} — prawdopodobnie jest otwarty "
            "w Excelu (albo w trakcie synchronizacji z chmurą). Zamknij go i spróbuj "
            f"ponownie. Szczegóły: {err}"
        ) from err


def _zapisz_workbook(wb: openpyxl.Workbook, sciezka: str) -> None:
    try:
        wb.save(sciezka)
    except OSError as err:
        raise BladNumeracji(
            f"Nie udało się zapisać pliku {sciezka} — prawdopodobnie jest otwarty "
            f"w Excelu. Zamknij go i spróbuj ponownie. Szczegóły: {err}"
        ) from err


def _sciezka_pliku() -> str:
    sciezka = ustawienia.wczytaj().get("sciezka_numery_zlecen")
    if not sciezka or not Path(sciezka).exists():
        raise BladNumeracji(
            "Nie ustawiono ścieżki do pliku Numery_zlecen_2026.xlsx (albo plik pod "
            "wskazaną ścieżką nie istnieje) — ustaw ją w Ustawieniach (ikona koła "
            "zębatego w prawym górnym rogu)."
        )
    return sciezka


def _arkusz(wb: openpyxl.Workbook, sciezka: str) -> Worksheet:
    if ARKUSZ not in wb.sheetnames:
        raise BladNumeracji(f"Plik {sciezka} nie ma zakładki „{ARKUSZ}”.")
    return wb[ARKUSZ]


def _znajdz_wiersz_numeru(ws: Worksheet, kol_numeru: str, numer: str) -> int | None:
    for r in range(2, ws.max_row + 1):
        if str(ws[f"{kol_numeru}{r}"].value or "").strip() == numer:
            return r
    return None


def zarezerwuj_numer(podmiot_realizujacy: str, account_manager: str | None = None) -> str:
    """Znajduje pierwszy wolny numer dla danego podmiotu (skan kolumny
    status), zapisuje w tym samym wierszu status "zajete", dzisiejszą datę i
    przekazanego account managera, po czym zwraca numer zlecenia (z kolumny
    numeru w tym samym wierszu)."""
    _sprawdz_podmiot(podmiot_realizujacy)
    kol = KOLUMNY[podmiot_realizujacy]
    sciezka = _sciezka_pliku()

    wb = _otworz_workbook(sciezka)
    try:
        ws = _arkusz(wb, sciezka)

        wiersz_wolny = None
        for r in range(2, ws.max_row + 1):
            status = ws[f"{kol['status']}{r}"].value
            if not status or not str(status).strip():
                wiersz_wolny = r
                break

        if wiersz_wolny is None:
            raise BladNumeracji(
                f"Brak wolnych numerów zleceń dla „{podmiot_realizujacy}” w pliku {sciezka}."
            )

        numer = ws[f"{kol['numer']}{wiersz_wolny}"].value
        if not numer:
            raise BladNumeracji(
                f"Wiersz {wiersz_wolny} w zakładce „{ARKUSZ}” nie ma numeru zlecenia "
                f"w kolumnie {kol['numer']}."
            )

        for klucz, wartosc in (
            ("status", "zajete"),
            ("data", date.today().strftime("%d.%m.%Y")),
            ("akant", account_manager or ""),
        ):
            komorka = ws[f"{kol[klucz]}{wiersz_wolny}"]
            komorka.value = wartosc
            komorka.alignment = WYROWNANIE_SRODEK

        _zapisz_workbook(wb, sciezka)
    finally:
        wb.close()

    return str(numer)


def zwolnij_numer(numer: str, podmiot_realizujacy: str) -> None:
    """Czyści rezerwację (status/data/akant) numeru wcześniej pobranego
    automatycznie — używane, gdy użytkownik na kroku 2 świadomie zachowuje
    numer wpisany ręcznie/wklejony z pliku kampanii zamiast tego pobranego
    automatycznie, żeby nie zostawić w pliku "wiszącej" rezerwacji numeru,
    którego nikt finalnie nie użył."""
    _sprawdz_podmiot(podmiot_realizujacy)
    kol = KOLUMNY[podmiot_realizujacy]
    sciezka = _sciezka_pliku()

    wb = _otworz_workbook(sciezka)
    try:
        ws = _arkusz(wb, sciezka)

        wiersz = _znajdz_wiersz_numeru(ws, kol["numer"], numer)
        if wiersz is None:
            raise BladNumeracji(
                f"Nie znaleziono numeru {numer} w pliku {sciezka} — rezerwacja nie została zwolniona."
            )

        for klucz in ("status", "data", "akant"):
            ws[f"{kol[klucz]}{wiersz}"] = None

        _zapisz_workbook(wb, sciezka)
    finally:
        wb.close()
