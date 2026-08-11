"""Generuje plik DANE_*.xlsx (dane dla traffic) — odpowiednik zakładki
"Dane Traffic" z app/data/DANE.xlsx (patrz źródła/Zlecenie/DANE.xlsx).

Zachowuje etykiety wierszy oryginalnego szablonu (poza "LP:", zawsze pustym
w rzeczywistych plikach — pominięte), w kolejności ustalonej z użytkownikiem
(numer/nazwa/kontakt na górze, żeby traffic od razu widział, co to za
zlecenie). Wiersze 16+ oryginału to gotowy pakiet aplikacji (App Title/Bundle
ID) dobrany pod grupę celową — świadomie pominięte na tym etapie (patrz
rozmowa z użytkownikiem: to ma docelowo wypluwać osobny system, nie ten
generator).

Nowość względem oryginału: tabela rozbicia liczby wyświetleń/klików na
miesiące (okresy) - oryginalny szablon miał tylko sumę całkowitą.

Formatowanie (ustalone z użytkownikiem): wszystko wyśrodkowane w pionie;
kolumna A (etykiety) do lewej w poziomie; wartości tekstowe wyśrodkowane w
poziomie; liczby do prawej w poziomie, bez miejsc po przecinku - poza
budżetami, które są kwotą w PLN z dwoma miejscami po przecinku. Daty liczą
się jak liczby przy wyrównaniu (tak jak domyślnie w Excelu), ale zachowują
własny format DD.MM.RRRR, nie "0 miejsc po przecinku".
"""
from datetime import date
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from openpyxl.worksheet.worksheet import Worksheet

from app.models.dane_traffic import DaneTraffic
from app.models.zlecenie import Zlecenie
from app.services.kalkulacje import MIESIACE_PL, etykieta_liczby, liczba_dla_okresu

FONT_ETYKIETA = Font(name="Calibri", size=11, bold=True)
FONT_NAGLOWEK = Font(name="Calibri", size=11, bold=True)

ALIGN_ETYKIETA = Alignment(horizontal="left", vertical="center")
ALIGN_TEKST = Alignment(horizontal="center", vertical="center")
ALIGN_LICZBA = Alignment(horizontal="right", vertical="center")

FORMAT_LICZBY = "#,##0"
FORMAT_KWOTY = '#,##0.00" PLN"'
FORMAT_DATY = "DD.MM.YYYY"


def _czy_liczba(wartosc) -> bool:
    return isinstance(wartosc, (int, float, date))


def _wiersz(ws: Worksheet, r: int, etykieta: str, wartosc) -> None:
    c_etykieta = ws.cell(row=r, column=1, value=etykieta)
    c_etykieta.font = FONT_ETYKIETA
    c_etykieta.alignment = ALIGN_ETYKIETA

    c_wartosc = ws.cell(row=r, column=2, value=wartosc)
    if isinstance(wartosc, date):
        c_wartosc.number_format = FORMAT_DATY
        c_wartosc.alignment = ALIGN_LICZBA
    elif _czy_liczba(wartosc):
        c_wartosc.number_format = FORMAT_LICZBY
        c_wartosc.alignment = ALIGN_LICZBA
    else:
        c_wartosc.alignment = ALIGN_TEKST


def _naglowek_tabeli(ws: Worksheet, r: int, teksty: list[str]) -> None:
    for c, tekst in enumerate(teksty, start=1):
        cell = ws.cell(row=r, column=c, value=tekst)
        cell.font = FONT_NAGLOWEK
        cell.alignment = ALIGN_ETYKIETA if c == 1 else ALIGN_TEKST


def _wiersz_tabeli(ws: Worksheet, r: int, miesiac: str, budzet: float, liczba: float, pogrubione: bool = False) -> None:
    # Kolejność kolumn (ustalona z użytkownikiem): B = Liczba (wyświetleń/
    # klików/...), C = Budżet - odwrotnie niż mogłoby się intuicyjnie
    # wydawać, ale tak ma zostać.
    czcionka = FONT_ETYKIETA if pogrubione else Font(name="Calibri", size=11)

    c_miesiac = ws.cell(row=r, column=1, value=miesiac)
    c_miesiac.font = czcionka
    c_miesiac.alignment = ALIGN_ETYKIETA

    c_liczba = ws.cell(row=r, column=2, value=liczba)
    c_liczba.font = czcionka
    c_liczba.number_format = FORMAT_LICZBY
    c_liczba.alignment = ALIGN_LICZBA

    c_budzet = ws.cell(row=r, column=3, value=budzet)
    c_budzet.font = czcionka
    c_budzet.number_format = FORMAT_KWOTY
    c_budzet.alignment = ALIGN_LICZBA


def _dopasuj_szerokosc(ws: Worksheet, kolumna: str, teksty: list[str], margines: int = 2) -> None:
    """Odpowiednik "auto-fit" szerokości kolumny w Excelu - openpyxl go nie
    ma wbudowanego, więc liczymy najdłuższy tekst faktycznie zapisany w tej
    kolumnie i dobieramy szerokość do niego, żeby nic się nie ucinało."""
    if not teksty:
        return
    ws.column_dimensions[kolumna].width = max(len(t) for t in teksty) + margines


def generuj_dane_traffic(zlecenie: Zlecenie, dane_traffic: DaneTraffic, sciezka: Path) -> Path:
    pola = zlecenie.pola
    capping_wartosc = pola.capping if pola.capping is not None else "brak"
    wydawcy_tekst = ", ".join(pola.wydawcy_zewnetrzni) if pola.wydawcy_zewnetrzni else "brak"

    wb = Workbook()
    ws = wb.active
    ws.title = "Dane Traffic"
    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 55

    r = 1
    for etykieta, wartosc in [
        ("Numer w moim pliku:", pola.nr_zlecenia),
        ("Nazwa kampanii:", pola.nazwa_kampanii),
        ("Osoba kontaktowa:", pola.zlecajacy),
        ("Model sprzedaży:", pola.model_sprzedazy),
        (f"{etykieta_liczby(pola.model_sprzedazy)}:", round(zlecenie.liczba_total)),
        ("Termin startu:", zlecenie.data_startu),
        ("Termin końca:", zlecenie.data_konca),
        ("Capp:", capping_wartosc),
        ("Formaty:", pola.format_reklamowy),
        ("Target:", pola.target),
        ("Spot:", dane_traffic.link_spot),
        ("Kody:", dane_traffic.link_kody),
        ("Uwagi dla traffic:", dane_traffic.uwagi_traffic),
        ("Spółka:", pola.podmiot_realizujacy),
        ("Wydawcy zewnętrzni:", wydawcy_tekst),
    ]:
        _wiersz(ws, r, etykieta, wartosc)
        r += 1

    r += 1  # pusty wiersz odstępu
    naglowek_sekcji = ws.cell(row=r, column=1, value="Rozbicie na miesiące")
    naglowek_sekcji.font = FONT_NAGLOWEK
    naglowek_sekcji.alignment = ALIGN_ETYKIETA
    r += 1

    etykieta_kolumny = etykieta_liczby(pola.model_sprzedazy)
    # Kolejność kolumn: B = Liczba, C = Budżet - patrz _wiersz_tabeli.
    _naglowek_tabeli(ws, r, ["Miesiąc", etykieta_kolumny, "Budżet"])
    r += 1
    teksty_kolumny_c = ["Budżet"]
    for okres in sorted(zlecenie.okresy, key=lambda o: o.data_startu):
        liczba = liczba_dla_okresu(pola.model_sprzedazy, pola.koszt_jednostkowy, okres.budzet)
        etykieta_miesiaca = f"{MIESIACE_PL[okres.data_startu.month - 1].capitalize()} {okres.data_startu.year}"
        budzet = round(okres.budzet, 2)
        _wiersz_tabeli(ws, r, etykieta_miesiaca, budzet, round(liczba))
        teksty_kolumny_c.append(f"{budzet:,.2f} PLN")
        r += 1

    budzet_total = round(zlecenie.budzet_total, 2)
    _wiersz_tabeli(ws, r, "Razem", budzet_total, round(zlecenie.liczba_total), pogrubione=True)
    teksty_kolumny_c.append(f"{budzet_total:,.2f} PLN")
    _dopasuj_szerokosc(ws, "C", teksty_kolumny_c)

    sciezka.parent.mkdir(parents=True, exist_ok=True)
    wb.save(sciezka)
    return sciezka
