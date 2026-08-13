from datetime import date

import openpyxl

from app.models.dane_traffic import DaneTraffic
from app.models.kampania import PolaWspolne
from app.models.okres import Okres
from app.models.zlecenie import Zlecenie
from app.services.generator_dane_traffic import generuj_dane_traffic


def _przykladowe_zlecenie(model_sprzedazy="CPM", koszt_jednostkowy=26, wydawcy_zewnetrzni=None):
    pola = PolaWspolne(
        account_manager="Igor Samul",
        podmiot_realizujacy="Sp. k.",
        nr_zlecenia="K/2026/077",
        nazwa_kampanii="Colian_Hellena_Lody_Cooltowe_Lipiec_Sierpien_Audio",
        dom_mediowy="Initiative Media Warszawa sp. z o.o.",
        klient="Colian",
        brand="Hellena",
        zlecajacy="Paulina Kowalik",
        target="KIDS",
        capping=3,
        format_reklamowy="In-game audio KIDS",
        model_sprzedazy=model_sprzedazy,
        koszt_jednostkowy=koszt_jednostkowy,
        uwagi="Dzieci",
        wydawcy_zewnetrzni=wydawcy_zewnetrzni or [],
    )
    okresy = [
        Okres(date(2026, 7, 28), date(2026, 7, 31), 5714.29),
        Okres(date(2026, 8, 1), date(2026, 8, 31), 44285.71),
    ]
    return Zlecenie(pola=pola, okresy=okresy)


def _wczytaj_pary_klucz_wartosc(sciezka):
    wb = openpyxl.load_workbook(sciezka)
    ws = wb.active
    wynik = {}
    for row in ws.iter_rows():
        klucz, wartosc = row[0].value, row[1].value
        if isinstance(klucz, str) and klucz.endswith(":"):
            wynik[klucz] = wartosc
    return wynik


def test_generuj_dane_traffic_pola_podstawowe(tmp_path):
    zlecenie = _przykladowe_zlecenie()
    dane_traffic = DaneTraffic(
        uwagi_traffic="Priorytet wysoki", link_spot="https://spot.example/1", link_kody="https://kody.example/1",
    )
    sciezka = generuj_dane_traffic(zlecenie, dane_traffic, tmp_path / "DANE_test.xlsx")
    assert sciezka.exists()

    pary = _wczytaj_pary_klucz_wartosc(sciezka)
    assert pary["Osoba kontaktowa:"] == "Paulina Kowalik"
    assert pary["Nazwa kampanii:"] == "Colian_Hellena_Lody_Cooltowe_Lipiec_Sierpien_Audio"
    assert pary["Numer w moim pliku:"] == "K/2026/077"
    assert pary["Model sprzedaży:"] == "CPM"
    assert pary["Capp:"] == 3
    assert pary["Formaty:"] == "In-game audio KIDS"
    assert pary["Target:"] == "KIDS"
    assert pary["Spot:"] == "https://spot.example/1"
    assert pary["Kody:"] == "https://kody.example/1"
    assert pary["Uwagi dla traffic:"] == "Priorytet wysoki"
    assert pary["Spółka:"] == "Sp. k."
    assert pary["Wydawcy zewnętrzni:"] == "brak"


def test_generuj_dane_traffic_kolejnosc_pierwszych_trzech_wierszy(tmp_path):
    # Ustalone z użytkownikiem: numer/nazwa/kontakt na górze, żeby traffic
    # od razu widział, co to za zlecenie, bez przewijania.
    zlecenie = _przykladowe_zlecenie()
    sciezka = generuj_dane_traffic(zlecenie, DaneTraffic(), tmp_path / "DANE_test.xlsx")
    ws = openpyxl.load_workbook(sciezka).active
    assert ws.cell(row=1, column=1).value == "Numer w moim pliku:"
    assert ws.cell(row=1, column=2).value == "K/2026/077"
    assert ws.cell(row=2, column=1).value == "Nazwa kampanii:"
    assert ws.cell(row=3, column=1).value == "Osoba kontaktowa:"


def test_generuj_dane_traffic_wyrownanie_i_formaty_liczb(tmp_path):
    zlecenie = _przykladowe_zlecenie()
    sciezka = generuj_dane_traffic(zlecenie, DaneTraffic(), tmp_path / "DANE_test.xlsx")
    ws = openpyxl.load_workbook(sciezka).active

    # Wszystko wyśrodkowane w pionie, kolumna A zawsze do lewej.
    for row in ws.iter_rows():
        for cell in row:
            if cell.value is None:
                continue
            assert cell.alignment.vertical == "center", f"{cell.coordinate}: {cell.value!r}"
            if cell.column == 1:
                assert cell.alignment.horizontal == "left", f"{cell.coordinate}: {cell.value!r}"

    pary_komorek = {row[0].value: row[1] for row in ws.iter_rows() if isinstance(row[0].value, str)}

    # Tekst w kolumnie B - wyśrodkowany.
    for klucz in ("Nazwa kampanii:", "Model sprzedaży:", "Spółka:"):
        assert pary_komorek[klucz].alignment.horizontal == "center", klucz

    # Liczby (w tym daty) w kolumnie B - do prawej.
    komorka_capp = pary_komorek["Capp:"]
    assert komorka_capp.alignment.horizontal == "right"
    assert komorka_capp.number_format == "#,##0"

    komorka_termin = pary_komorek["Termin startu:"]
    assert komorka_termin.alignment.horizontal == "right"
    assert komorka_termin.number_format == "DD.MM.YYYY"

    komorka_liczba = pary_komorek["Liczba wyświetleń:"]
    assert komorka_liczba.alignment.horizontal == "right"
    assert komorka_liczba.number_format == "#,##0"

    # Tabela miesięczna: kolumna B = Liczba, kolumna C = Budżet (waluta PLN,
    # dwa miejsca po przecinku) - ustalona z użytkownikiem kolejność.
    for row in ws.iter_rows():
        if row[0].value in ("Lipiec 2026", "Sierpień 2026", "Razem"):
            komorka_liczba_miesiac = row[1]
            assert komorka_liczba_miesiac.number_format == "#,##0"
            assert komorka_liczba_miesiac.alignment.horizontal == "right"
            komorka_budzet = row[2]
            assert komorka_budzet.number_format == '#,##0.00" PLN"'
            assert komorka_budzet.alignment.horizontal == "right"


def test_generuj_dane_traffic_wydawca_zewnetrzny(tmp_path):
    zlecenie = _przykladowe_zlecenie(wydawcy_zewnetrzni=["KIDOZ"])
    sciezka = generuj_dane_traffic(zlecenie, DaneTraffic(), tmp_path / "DANE_test.xlsx")
    pary = _wczytaj_pary_klucz_wartosc(sciezka)
    assert pary["Wydawcy zewnętrzni:"] == "KIDOZ"


def test_generuj_dane_traffic_wieloliniowe_uwagi_maja_wrap_text_i_wysokosc(tmp_path):
    # Regresja: bez wrap_text Excel ignoruje \n w komórce i sklein wszystko
    # w jeden ciąg bez żadnego odstępu (zgłoszone przez użytkownika).
    zlecenie = _przykladowe_zlecenie()
    tekst = "Linia pierwsza\nLinia druga\nLinia trzecia"
    dane_traffic = DaneTraffic(uwagi_traffic=tekst)
    sciezka = generuj_dane_traffic(zlecenie, dane_traffic, tmp_path / "DANE_test.xlsx")

    ws = openpyxl.load_workbook(sciezka).active
    for row in ws.iter_rows():
        if row[0].value == "Uwagi dla traffic:":
            komorka = row[1]
            assert komorka.value == tekst
            assert komorka.alignment.wrap_text is True
            assert ws.row_dimensions[komorka.row].height == 3 * 15.0
            break
    else:
        raise AssertionError("nie znaleziono wiersza 'Uwagi dla traffic:'")


def test_generuj_dane_traffic_bez_lp_wiersza(tmp_path):
    # "LP:" z oryginalnego szablonu jest zawsze puste w praktyce - świadomie
    # pominięte przy generowaniu (patrz docstring generator_dane_traffic.py).
    zlecenie = _przykladowe_zlecenie()
    sciezka = generuj_dane_traffic(zlecenie, DaneTraffic(), tmp_path / "DANE_test.xlsx")
    pary = _wczytaj_pary_klucz_wartosc(sciezka)
    assert "LP:" not in pary


def test_generuj_dane_traffic_kolejnosc_kolumn_tabeli_miesiecznej(tmp_path):
    # Ustalone z użytkownikiem: B = Liczba, C = Budżet (nie odwrotnie).
    zlecenie = _przykladowe_zlecenie()
    sciezka = generuj_dane_traffic(zlecenie, DaneTraffic(), tmp_path / "DANE_test.xlsx")
    ws = openpyxl.load_workbook(sciezka).active
    for row in ws.iter_rows():
        if row[0].value == "Miesiąc":
            assert row[1].value == "Liczba wyświetleń"
            assert row[2].value == "Budżet"
            break
    else:
        raise AssertionError("nie znaleziono wiersza nagłówka tabeli miesięcznej")


def test_generuj_dane_traffic_kolumna_c_ma_dopasowana_szerokosc(tmp_path):
    zlecenie = _przykladowe_zlecenie()
    sciezka = generuj_dane_traffic(zlecenie, DaneTraffic(), tmp_path / "DANE_test.xlsx")
    ws = openpyxl.load_workbook(sciezka).active
    najdluzszy = max(
        len("Budżet"),
        *(len(f"{round(o.budzet, 2):,.2f} PLN") for o in zlecenie.okresy),
        len(f"{round(zlecenie.budzet_total, 2):,.2f} PLN"),
    )
    assert ws.column_dimensions["C"].width == najdluzszy + 2


def test_generuj_dane_traffic_rozbicie_na_miesiace(tmp_path):
    zlecenie = _przykladowe_zlecenie()
    sciezka = generuj_dane_traffic(zlecenie, DaneTraffic(), tmp_path / "DANE_test.xlsx")

    wb = openpyxl.load_workbook(sciezka)
    ws = wb.active
    tekst_calego_arkusza = [str(c.value) for row in ws.iter_rows() for c in row if c.value is not None]
    assert "Rozbicie na miesiące" in tekst_calego_arkusza
    assert "Lipiec 2026" in tekst_calego_arkusza
    assert "Sierpień 2026" in tekst_calego_arkusza
    assert "Razem" in tekst_calego_arkusza
