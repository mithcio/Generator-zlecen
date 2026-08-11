"""Testy warstwy UI bez żywego Page/przeglądarki: buduj() tylko konstruuje
drzewo kontrolek Flet (ft.Dropdown/TextField/...), więc można je wywołać
bezpośrednio i sprawdzić, że się nie wysypie oraz że handlery przycisków
działają na aktualnym stanie, a nie na zamrożonym w chwili renderu."""
import json
from datetime import date
from pathlib import Path

import flet as ft
import openpyxl
import pytest

from app.models.okres import Okres
from app.services import cennik, ustawienia
from app.ui import (
    krok1_podmiot,
    krok2_dane_kampanii,
    krok3_okresy,
    krok4_podglad,
    krok5_dane_traffic,
)
from app.ui.stan import StanKreatora

# PELNY_STAN (niżej) używa zawsze format_reklamowy="In-game audio KIDS" i
# target="KIDS" - stawki poniżej odpowiadają kluczom, jakie z tego wynikają
# (patrz generator_wydawcy._kidoz_typ_kampanii/_adverty_placement/
# _kidoz_wydawca_cennika), mimo że biznesowo nie każdy wydawca faktycznie
# obsługuje ten format (nieistotne dla testów UI).
CENNIK_TESTOWY = {
    "KIDOZ": {"Rewarded": {"cena": 2.5, "waluta": "USD"}},
    "Adverty": {"Video": {"cena": 3.5, "waluta": "USD"}},
    "Odeeo": {"In-Game Audio": {"cena": 2.0, "waluta": "EUR"}},
    "Crazygames": {"Video": {"cena": 2.5, "waluta": "EUR"}},
    "POKI": {"ImViTa": {"cena": 3.0, "waluta": "EUR"}},
}


def _zbuduj_plik_numery(tmp_path, ile=50):
    """Kopia struktury Numery_zlecen_2026.xlsx z komplet wolnych numerów -
    krok2 rezerwuje numer naprawdę (zapisuje na dysk), więc testy potrzebują
    własnego pliku zamiast dotykania prawdziwego rejestru."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Numery zleceń"
    ws.append(["SP. K.", "status", "data", "użytkownik", None, "SP. Z O.O.", "status", "data", "użytkownik"])
    for i in range(1, ile + 1):
        ws.append([f"K/2026/{i:03d}", None, None, None, None, f"S/2026/{i:03d}", None, None, None])
    plik = tmp_path / "Numery_zlecen_2026.xlsx"
    wb.save(plik)
    return plik


@pytest.fixture(autouse=True)
def _izolacja_plikow_zewnetrznych(tmp_path, monkeypatch):
    """Izoluje wszystko, co app/services/numeracja.py i eksport_nazwy.py
    normalnie czytają z app/data/ustawienia.json — bez tego testy zależałyby
    od (i mogłyby nadpisać) prawdziwe ustawienia na maszynie dewelopera."""
    monkeypatch.setattr(ustawienia, "DATA_PLIK", tmp_path / "ustawienia.json")
    ustawienia.zapisz(sciezka_numery_zlecen=str(_zbuduj_plik_numery(tmp_path)))

    plik_cennika = tmp_path / "cennik_wydawcow.json"
    plik_cennika.write_text(json.dumps(CENNIK_TESTOWY), encoding="utf-8")
    monkeypatch.setattr(cennik, "DATA_PLIK", plik_cennika)


class FakePage:
    """Wystarczy do przechwycenia AlertDialog wołanego przez kreator.page.*
    (dialogi konfliktów/ostrzeżeń w krok2) bez prawdziwego Flet Page."""

    def __init__(self):
        self.dialog: ft.AlertDialog | None = None

    def show_dialog(self, dlg):
        self.dialog = dlg

    def pop_dialog(self):
        self.dialog = None


class FakeKreator:
    def __init__(self, stan):
        self.stan = stan
        self.page = FakePage()
        self.bledy_pokazane: list = []
        self.liczba_odswiezen = 0

    def odswiez(self):
        self.liczba_odswiezen += 1

    def wroc(self):
        if self.stan.krok > 1:
            self.stan.krok -= 1

    def idz_do_kroku(self, krok):
        self.stan.krok = krok

    def pokaz_blad(self, bledy):
        self.bledy_pokazane.append(bledy)

    def pokaz_ustawienia(self):
        pass


def _znajdz_przez_tekst(control, klasa, tekst):
    if control is None:
        return None
    if isinstance(control, klasa) and getattr(control, "content", None) == tekst:
        return control
    for c in list(getattr(control, "controls", []) or []) + list(getattr(control, "actions", []) or []):
        wynik = _znajdz_przez_tekst(c, klasa, tekst)
        if wynik:
            return wynik
    content = getattr(control, "content", None)
    if isinstance(content, ft.Control):
        wynik = _znajdz_przez_tekst(content, klasa, tekst)
        if wynik:
            return wynik
    return None


def _znajdz_dropdown_po_etykiecie(control, etykieta):
    if control is None:
        return None
    if isinstance(control, ft.Dropdown) and control.label == etykieta:
        return control
    for c in list(getattr(control, "controls", []) or []) + list(getattr(control, "actions", []) or []):
        wynik = _znajdz_dropdown_po_etykiecie(c, etykieta)
        if wynik:
            return wynik
    content = getattr(control, "content", None)
    if isinstance(content, ft.Control):
        return _znajdz_dropdown_po_etykiecie(content, etykieta)
    return None


def _znajdz_wszystkie_ikony(control, ikona):
    wyniki = []
    if isinstance(control, ft.IconButton) and control.icon == ikona:
        wyniki.append(control)
    for c in getattr(control, "controls", []) or []:
        wyniki.extend(_znajdz_wszystkie_ikony(c, ikona))
    content = getattr(control, "content", None)
    if isinstance(content, ft.Control):
        wyniki.extend(_znajdz_wszystkie_ikony(content, ikona))
    return wyniki


PELNY_STAN = dict(
    account_manager="Igor Samul",
    podmiot_realizujacy="Sp. k.",
    nazwa_kampanii="Test",
    dom_mediowy="Initiative Media Warszawa sp. z o.o.",
    klient="Colian",
    brand="Hellena",
    zlecajacy="Paulina Kowalik",
    target="KIDS",
    capping="3",
    format_reklamowy="In-game audio KIDS",
    model_sprzedazy="CPM",
    koszt_jednostkowy="26",
    nr_zlecenia="K/2026/078",
)


@pytest.mark.parametrize(
    "modul, stan_kwargs",
    [
        (krok1_podmiot, {}),
        (krok1_podmiot, dict(account_manager="Igor Samul", podmiot_realizujacy="Sp. k.")),
        (krok2_dane_kampanii, {}),  # regresja: skok tu bez wypełnienia kroku 1 walił KeyError
        (krok2_dane_kampanii, dict(account_manager="Igor Samul", podmiot_realizujacy="Sp. k.", tryb_danych="formularz")),
        (krok2_dane_kampanii, dict(account_manager="Igor Samul", podmiot_realizujacy="Sp. k.", tryb_danych="wklej")),
        (krok2_dane_kampanii, {**PELNY_STAN, "tryb_danych": "formularz"}),
        (krok2_dane_kampanii, dict(account_manager="Igor Samul", podmiot_realizujacy="Sp. z o.o.", tryb_danych="formularz")),
        (
            krok2_dane_kampanii,
            {**PELNY_STAN, "podmiot_realizujacy": "Sp. z o.o.", "tryb_danych": "formularz"},
        ),
        (krok3_okresy, dict(account_manager="Igor Samul", podmiot_realizujacy="Sp. k.")),
        (
            krok3_okresy,
            dict(
                account_manager="Igor Samul",
                podmiot_realizujacy="Sp. k.",
                okresy=[Okres(date(2026, 7, 28), date(2026, 7, 31), 5714.29)],
            ),
        ),
        (
            krok3_okresy,
            dict(
                account_manager="Igor Samul", podmiot_realizujacy="Sp. k.",
                kalendarz_pole="start", kalendarz_rok=2026, kalendarz_miesiac=7,
            ),
        ),
        (
            krok3_okresy,
            dict(
                account_manager="Igor Samul", podmiot_realizujacy="Sp. k.",
                kalendarz_pole="koniec", kalendarz_rok=2026, kalendarz_miesiac=7,
                nowy_okres_start=date(2026, 7, 15),
            ),
        ),
        (krok4_podglad, dict(account_manager="Igor Samul", podmiot_realizujacy="Sp. k.")),
        (krok4_podglad, {**PELNY_STAN, "okresy": [Okres(date(2026, 7, 28), date(2026, 7, 31), 5714.29)]}),
        (
            krok4_podglad,
            {**{**PELNY_STAN, "dom_mediowy": "Nieznana Agencja"}, "okresy": [Okres(date(2026, 7, 28), date(2026, 7, 31), 5714.29)]},
        ),
        (
            krok4_podglad,
            {
                **PELNY_STAN,
                "okresy": [Okres(date(2026, 7, 28), date(2026, 7, 31), 5714.29)],
                "pokaz_wiersze_kampanii": True,
            },
        ),
        (
            krok4_podglad,
            {
                **PELNY_STAN,
                "okresy": [Okres(date(2026, 7, 28), date(2026, 7, 31), 5714.29)],
                "zlecenie_wygenerowane": ("C:/folder", "C:/folder/plik.xlsx", "C:/folder/plik.pdf"),
            },
        ),
        (krok5_dane_traffic, dict(account_manager="Igor Samul", podmiot_realizujacy="Sp. k.")),
        (krok5_dane_traffic, {**PELNY_STAN, "okresy": [Okres(date(2026, 7, 28), date(2026, 7, 31), 5714.29)]}),
        (
            krok5_dane_traffic,
            {
                **PELNY_STAN,
                "okresy": [
                    Okres(date(2026, 7, 1), date(2026, 7, 31), 1000.0),
                    Okres(date(2026, 8, 1), date(2026, 8, 31), 2000.0),
                ],
                "wydawca_zewnetrzny": "KIDOZ",
            },
        ),
    ],
)
def test_buduj_nie_wybucha(modul, stan_kwargs):
    stan = StanKreatora(**stan_kwargs)
    kontrolka = modul.buduj(FakeKreator(stan))
    assert isinstance(kontrolka, ft.Control)


def test_krok2_bez_kroku1_pokazuje_komunikat_zamiast_crashowac():
    stan = StanKreatora(krok=2)  # account_manager i podmiot_realizujacy puste
    kreator = FakeKreator(stan)
    kontrolka = krok2_dane_kampanii.buduj(kreator)
    przycisk = _znajdz_przez_tekst(kontrolka, ft.FilledButton, "Wróć do kroku 1")
    assert przycisk is not None
    przycisk.on_click(None)
    assert stan.krok == 1


def _znajdz_pola_tylko_do_odczytu(control, wynik=None):
    if wynik is None:
        wynik = []
    if isinstance(control, ft.TextField) and control.read_only:
        wynik.append(control)
    for c in getattr(control, "controls", []) or []:
        _znajdz_pola_tylko_do_odczytu(c, wynik)
    content = getattr(control, "content", None)
    if isinstance(content, ft.Control):
        _znajdz_pola_tylko_do_odczytu(content, wynik)
    return wynik


def test_krok4_przycisk_pokazuje_wiersze_bez_generowania_zlecenia():
    # Wiersze do pliku kampanii mają być dostępne w kroku 4 bez konieczności
    # generowania xlsx/PDF.
    stan = StanKreatora(
        krok=4, **PELNY_STAN,
        okresy=[
            Okres(date(2026, 7, 1), date(2026, 7, 31), 1000.0),
            Okres(date(2026, 8, 1), date(2026, 8, 31), 2000.0),
        ],
    )
    kreator = FakeKreator(stan)
    kontrolka = krok4_podglad.buduj(kreator)
    assert not _znajdz_pola_tylko_do_odczytu(kontrolka)

    przycisk = _znajdz_przez_tekst(kontrolka, ft.OutlinedButton, "Pokaż wiersz(e) do pliku kampanii")
    assert przycisk is not None
    przycisk.on_click(None)
    assert stan.pokaz_wiersze_kampanii is True
    assert stan.zlecenie_wygenerowane is None  # nic nie zostało wygenerowane

    kontrolka = krok4_podglad.buduj(FakeKreator(stan))
    pola = _znajdz_pola_tylko_do_odczytu(kontrolka)
    assert len(pola) == 2  # jedno na okres


def test_krok4_generuj_zlecenie_zapisuje_pliki_i_pokazuje_wynik(tmp_path):
    ustawienia.zapisz(folder_eksportu=str(tmp_path))
    stan = StanKreatora(
        krok=4, **PELNY_STAN,
        okresy=[Okres(date(2026, 7, 28), date(2026, 7, 31), 5714.29)],
    )
    kreator = FakeKreator(stan)
    kontrolka = krok4_podglad.buduj(kreator)
    assert stan.zlecenie_wygenerowane is None

    przycisk = _znajdz_przez_tekst(kontrolka, ft.FilledButton, "Generuj zlecenie")
    assert przycisk is not None
    przycisk.on_click(None)

    assert stan.zlecenie_wygenerowane is not None
    folder, xlsx, pdf = stan.zlecenie_wygenerowane
    assert Path(xlsx).exists()
    assert Path(pdf).exists()
    assert stan.krok == 4  # zostaje na miejscu, nie skacze dalej samo z siebie


def _znajdz_tekst_po_wartosci(control, wartosc):
    """ft.Text trzyma swój napis w .value (nie .content jak przyciski) -
    _znajdz_przez_tekst tu nie pasuje."""
    if isinstance(control, ft.Text) and control.value == wartosc:
        return control
    for c in getattr(control, "controls", []) or []:
        wynik = _znajdz_tekst_po_wartosci(c, wartosc)
        if wynik:
            return wynik
    content = getattr(control, "content", None)
    if isinstance(content, ft.Control):
        return _znajdz_tekst_po_wartosci(content, wartosc)
    return None


def test_krok5_traffic_tabela_pokazuje_rozbicie_na_miesiace():
    stan = StanKreatora(
        krok=5, **PELNY_STAN,
        okresy=[
            Okres(date(2026, 7, 1), date(2026, 7, 31), 10000.0),
            Okres(date(2026, 8, 1), date(2026, 8, 31), 20000.0),
        ],
    )
    kontrolka = krok5_dane_traffic.buduj(FakeKreator(stan))
    assert _znajdz_tekst_po_wartosci(kontrolka, "Rozbicie na miesiące") is not None
    assert _znajdz_tekst_po_wartosci(kontrolka, "Lipiec 2026") is not None
    assert _znajdz_tekst_po_wartosci(kontrolka, "Sierpień 2026") is not None
    assert _znajdz_tekst_po_wartosci(kontrolka, "Razem") is not None


def test_krok5_traffic_pola_edytowalne_aktualizuja_stan():
    stan = StanKreatora(
        krok=5, **PELNY_STAN,
        okresy=[Okres(date(2026, 7, 1), date(2026, 7, 31), 1000.0)],
    )
    kontrolka = krok5_dane_traffic.buduj(FakeKreator(stan))

    for etykieta, pole_stanu, wartosc in [
        ("Uwagi dla traffic", "uwagi_traffic", "Priorytet wysoki"),
        ("Link do Spotu", "link_spot", "https://spot.example/1"),
        ("Link do Kodów", "link_kody", "https://kody.example/1"),
    ]:
        pole = _znajdz_pole_tekstowe_po_etykiecie(kontrolka, etykieta)
        assert pole is not None, etykieta

        class FakeEvent:
            control = type("C", (), {"value": wartosc})()

        pole.on_change(FakeEvent())
        assert getattr(stan, pole_stanu) == wartosc


def test_krok5_traffic_wydawca_dropdown_aktualizuje_stan():
    stan = StanKreatora(
        krok=5, **PELNY_STAN,
        okresy=[Okres(date(2026, 7, 1), date(2026, 7, 31), 1000.0)],
    )
    kreator = FakeKreator(stan)
    kontrolka = krok5_dane_traffic.buduj(kreator)
    dropdown = _znajdz_dropdown_po_etykiecie(kontrolka, "Wydawcy zewnętrzni")
    assert dropdown is not None

    class FakeEvent:
        control = type("C", (), {"value": "KIDOZ"})()

    dropdown.on_select(FakeEvent())
    assert stan.wydawca_zewnetrzny == "KIDOZ"
    assert kreator.liczba_odswiezen == 1


def test_krok5_traffic_nowe_zlecenie_resetuje_stan():
    stan = StanKreatora(
        krok=5, **PELNY_STAN,
        okresy=[Okres(date(2026, 7, 1), date(2026, 7, 31), 1000.0)],
    )
    kreator = FakeKreator(stan)
    kontrolka = krok5_dane_traffic.buduj(kreator)
    _znajdz_przez_tekst(kontrolka, ft.FilledButton, "Nowe zlecenie").on_click(None)
    assert kreator.stan is not stan
    assert kreator.stan.krok == 1
    assert kreator.stan.nr_zlecenia == ""


def test_krok5_traffic_generuje_sam_plik_bez_przechodzenia_dalej(tmp_path):
    # Dane Traffic generuje się niezależnie od Zlecenie (krok 4) - osobny
    # plik, osobny przycisk.
    ustawienia.zapisz(folder_eksportu=str(tmp_path))
    stan = StanKreatora(
        krok=5, **PELNY_STAN,
        okresy=[Okres(date(2026, 7, 1), date(2026, 7, 31), 1000.0)],
    )
    kreator = FakeKreator(stan)
    kontrolka = krok5_dane_traffic.buduj(kreator)
    assert stan.ostatnia_sciezka_dane_traffic is None

    przycisk = _znajdz_przez_tekst(kontrolka, ft.OutlinedButton, "Generuj plik Dane Traffic")
    assert przycisk is not None
    przycisk.on_click(None)

    assert stan.ostatnia_sciezka_dane_traffic is not None
    assert Path(stan.ostatnia_sciezka_dane_traffic).exists()
    assert stan.krok == 5  # zostaje na miejscu
    assert stan.zlecenie_wygenerowane is None  # nie dotyka generowania Zlecenie (krok 4)
    assert stan.ostatnia_sciezka_wydawcy is None  # wydawca_zewnetrzny domyślnie "brak"


def test_krok5_traffic_z_wydawcem_generuje_dodatkowy_plik(tmp_path):
    ustawienia.zapisz(folder_eksportu=str(tmp_path))
    stan = StanKreatora(
        krok=5, **{**PELNY_STAN, "format_reklamowy": "Rewarded KIDS"}, wydawca_zewnetrzny="KIDOZ",
        okresy=[Okres(date(2026, 7, 1), date(2026, 7, 31), 1000.0)],
    )
    kreator = FakeKreator(stan)
    kontrolka = krok5_dane_traffic.buduj(kreator)

    przycisk = _znajdz_przez_tekst(kontrolka, ft.OutlinedButton, "Generuj Dane Traffic + plik KIDOZ")
    assert przycisk is not None
    przycisk.on_click(None)

    assert stan.ostatnia_sciezka_dane_traffic is not None
    assert Path(stan.ostatnia_sciezka_dane_traffic).exists()
    assert stan.ostatnia_sciezka_wydawcy is not None
    assert Path(stan.ostatnia_sciezka_wydawcy).exists()
    assert "KIDOZ" in stan.ostatnia_sciezka_wydawcy


def test_krok5_traffic_poki_pokazuje_dropdown_placementu():
    stan = StanKreatora(
        krok=5, **PELNY_STAN, wydawca_zewnetrzny="POKI",
        okresy=[Okres(date(2026, 7, 1), date(2026, 7, 31), 1000.0)],
    )
    kontrolka = krok5_dane_traffic.buduj(FakeKreator(stan))
    assert _znajdz_dropdown_po_etykiecie(kontrolka, "Placement POKI") is not None


def test_krok5_traffic_bez_poki_nie_pokazuje_dropdown_placementu():
    stan = StanKreatora(
        krok=5, **PELNY_STAN, wydawca_zewnetrzny="KIDOZ",
        okresy=[Okres(date(2026, 7, 1), date(2026, 7, 31), 1000.0)],
    )
    kontrolka = krok5_dane_traffic.buduj(FakeKreator(stan))
    assert _znajdz_dropdown_po_etykiecie(kontrolka, "Placement POKI") is None


def test_krok5_traffic_poki_bez_placementu_pokazuje_blad(tmp_path):
    ustawienia.zapisz(folder_eksportu=str(tmp_path))
    stan = StanKreatora(
        krok=5, **{**PELNY_STAN, "format_reklamowy": "Rewarded KIDS"}, wydawca_zewnetrzny="POKI",
        okresy=[Okres(date(2026, 7, 1), date(2026, 7, 31), 1000.0)],
    )
    kreator = FakeKreator(stan)
    kontrolka = krok5_dane_traffic.buduj(kreator)
    przycisk = _znajdz_przez_tekst(kontrolka, ft.OutlinedButton, "Generuj Dane Traffic + plik POKI")
    przycisk.on_click(None)
    assert stan.ostatnia_sciezka_wydawcy is None
    assert kreator.bledy_pokazane != []


def test_krok5_traffic_poki_z_placementem_generuje_plik(tmp_path):
    ustawienia.zapisz(folder_eksportu=str(tmp_path))
    stan = StanKreatora(
        krok=5, **{**PELNY_STAN, "format_reklamowy": "Rewarded KIDS"},
        wydawca_zewnetrzny="POKI", poki_placement="ImViTa",
        okresy=[Okres(date(2026, 7, 1), date(2026, 7, 31), 1000.0)],
    )
    kreator = FakeKreator(stan)
    kontrolka = krok5_dane_traffic.buduj(kreator)
    przycisk = _znajdz_przez_tekst(kontrolka, ft.OutlinedButton, "Generuj Dane Traffic + plik POKI")
    przycisk.on_click(None)
    assert stan.ostatnia_sciezka_wydawcy is not None
    assert Path(stan.ostatnia_sciezka_wydawcy).exists()
    assert "POKI" in stan.ostatnia_sciezka_wydawcy


def test_krok5_traffic_format_niezgodny_z_wydawca_pokazuje_ostrzezenie(tmp_path):
    # In-game audio pasuje do Odeeo, nie do KIDOZ.
    ustawienia.zapisz(folder_eksportu=str(tmp_path))
    stan = StanKreatora(
        krok=5, **{**PELNY_STAN, "format_reklamowy": "In-game audio KIDS"}, wydawca_zewnetrzny="KIDOZ",
        okresy=[Okres(date(2026, 7, 1), date(2026, 7, 31), 1000.0)],
    )
    kreator = FakeKreator(stan)
    kontrolka = krok5_dane_traffic.buduj(kreator)
    przycisk = _znajdz_przez_tekst(kontrolka, ft.OutlinedButton, "Generuj Dane Traffic + plik KIDOZ")
    przycisk.on_click(None)

    assert stan.ostatnia_sciezka_dane_traffic is None  # nic nie wygenerowane - czeka na decyzję
    assert kreator.page.dialog is not None

    przycisk_kontynuuj = _znajdz_przez_tekst(kreator.page.dialog, ft.FilledButton, "Kontynuuj mimo to")
    assert przycisk_kontynuuj is not None
    przycisk_kontynuuj.on_click(None)

    assert stan.ostatnia_sciezka_dane_traffic is not None
    assert stan.ostatnia_sciezka_wydawcy is not None


def test_krok5_traffic_format_zgodny_z_wydawca_bez_ostrzezenia(tmp_path):
    ustawienia.zapisz(folder_eksportu=str(tmp_path))
    stan = StanKreatora(
        krok=5, **{**PELNY_STAN, "format_reklamowy": "In-game audio KIDS"}, wydawca_zewnetrzny="ODEEO",
        okresy=[Okres(date(2026, 7, 1), date(2026, 7, 31), 1000.0)],
    )
    kreator = FakeKreator(stan)
    kontrolka = krok5_dane_traffic.buduj(kreator)
    przycisk = _znajdz_przez_tekst(kontrolka, ft.OutlinedButton, "Generuj Dane Traffic + plik ODEEO")
    przycisk.on_click(None)

    assert kreator.page.dialog is None
    assert stan.ostatnia_sciezka_wydawcy is not None


def test_krok1_bez_domyslnego_konta_dropdown_odblokowany():
    stan = StanKreatora()
    kontrolka = krok1_podmiot.buduj(FakeKreator(stan))
    dropdown = _znajdz_dropdown_po_etykiecie(kontrolka, "Account manager")
    assert dropdown is not None
    assert dropdown.disabled is False


def test_krok1_domyslny_account_blokuje_dropdown_i_wymusza_wartosc():
    ustawienia.zapisz(domyslny_account_manager="Igor Samul")
    stan = StanKreatora()
    kontrolka = krok1_podmiot.buduj(FakeKreator(stan))
    dropdown = _znajdz_dropdown_po_etykiecie(kontrolka, "Account manager")
    assert dropdown is not None
    assert dropdown.disabled is True
    assert dropdown.value == "Igor Samul"
    assert stan.account_manager == "Igor Samul"


def test_krok1_domyslny_brak_nie_blokuje_dropdown():
    ustawienia.zapisz(domyslny_account_manager=None)
    stan = StanKreatora()
    kontrolka = krok1_podmiot.buduj(FakeKreator(stan))
    dropdown = _znajdz_dropdown_po_etykiecie(kontrolka, "Account manager")
    assert dropdown.disabled is False


def test_krok2_formularz_spk_ma_dom_mediowy_i_osobny_klient():
    """Sp. k. (agencja pośredniczy): "DOM Mediowy" (agencje) i osobne
    "Klient" (marki obsługiwane przez wybraną agencję) to dwa różne pola."""
    stan = StanKreatora(account_manager="Igor Samul", podmiot_realizujacy="Sp. k.", tryb_danych="formularz")
    kontrolka = krok2_dane_kampanii.buduj(FakeKreator(stan))
    assert _znajdz_dropdown_po_etykiecie(kontrolka, "DOM Mediowy") is not None
    assert _znajdz_dropdown_po_etykiecie(kontrolka, "Klient") is not None


def test_krok2_formularz_spzoo_ma_tylko_jedno_pole_klient():
    """Sp. z o.o. (klient bezpośredni): nie ma osobnej agencji - "DOM
    Mediowy" znika, zostaje jedno pole "Klient" (bez zdublowanego)."""
    stan = StanKreatora(account_manager="Igor Samul", podmiot_realizujacy="Sp. z o.o.", tryb_danych="formularz")
    kontrolka = krok2_dane_kampanii.buduj(FakeKreator(stan))
    assert _znajdz_dropdown_po_etykiecie(kontrolka, "DOM Mediowy") is None
    assert _znajdz_dropdown_po_etykiecie(kontrolka, "Klient") is not None


def test_krok2_formularz_spzoo_wybor_klienta_synchronizuje_dom_mediowy_i_klient():
    stan = StanKreatora(account_manager="Igor Samul", podmiot_realizujacy="Sp. z o.o.", tryb_danych="formularz")
    kontrolka = krok2_dane_kampanii.buduj(FakeKreator(stan))
    pole_klient = _znajdz_dropdown_po_etykiecie(kontrolka, "Klient")
    assert pole_klient is not None

    class FakeEvent:
        control = type("C", (), {"value": "TM Toys sp. z o.o."})()

    pole_klient.on_select(FakeEvent())
    assert stan.dom_mediowy == "TM Toys sp. z o.o."
    assert stan.klient == "TM Toys sp. z o.o."


def test_krok2_formularz_spk_wybor_agencji_resetuje_klienta_i_odswieza():
    stan = StanKreatora(account_manager="Igor Samul", podmiot_realizujacy="Sp. k.", klient="Stare Cos", tryb_danych="formularz")
    kreator = FakeKreator(stan)
    kontrolka = krok2_dane_kampanii.buduj(kreator)
    pole_agencja = _znajdz_dropdown_po_etykiecie(kontrolka, "DOM Mediowy")
    assert pole_agencja is not None

    class FakeEvent:
        control = type("C", (), {"value": "Starcom Sp. z o.o."})()

    pole_agencja.on_select(FakeEvent())
    assert stan.dom_mediowy == "Starcom Sp. z o.o."
    assert stan.klient == ""
    assert kreator.liczba_odswiezen == 1


def test_krok2_wklej_spzoo_ignoruje_kolumne_klient():
    """Wiersz klienta bezpośredniego z "-" w kolumnie Klient - ma być
    zignorowana, stan.klient ma się zsynchronizować z dom_mediowy."""
    wiersz = "\t".join(
        [
            "Test", "TM Toys sp. z o.o.", "-", "Jan Testowy",
            "KIDS", "NIE", "In-game audio KIDS", "Sp. z o.o.", "", "CPM", "26",
            "S/2026/078", "1000", "01.07.2026", "31.07.2026",
        ]
    )
    stan = StanKreatora(
        krok=2, account_manager="Igor Samul", podmiot_realizujacy="Sp. z o.o.", tryb_danych="wklej",
        wiersze_wklejane=[wiersz],
    )
    kreator = FakeKreator(stan)
    kontrolka = krok2_dane_kampanii.buduj(kreator)
    przycisk = _znajdz_przez_tekst(kontrolka, ft.FilledButton, "Wczytaj wiersze")
    przycisk.on_click(None)

    assert stan.dom_mediowy == "TM Toys sp. z o.o."
    assert stan.klient == "TM Toys sp. z o.o."  # nie "-"


def test_krok2_wklej_nie_ustawia_uwag_z_kolumny_uwagi_pliku_kampanii():
    # Kolumna Uwagi w pliku kampanii to co innego niż Zlecenie.pola.uwagi
    # (uwaga na dokumencie dla klienta) - nigdy nie ma trafiać do stanu
    # automatycznie, tylko ręcznie w formularzu.
    wiersz = "\t".join(
        [
            "Test", "Initiative Media Warszawa sp. z o.o.", "Colian", "Paulina Kowalik",
            "KIDS", "NIE", "In-game audio KIDS", "Sp. k.", "Jakaś uwaga z pliku kampanii", "CPM", "26",
            "K/2026/078", "1000", "01.07.2026", "31.07.2026",
        ]
    )
    stan = StanKreatora(
        krok=2, account_manager="Igor Samul", podmiot_realizujacy="Sp. k.", tryb_danych="wklej",
        wiersze_wklejane=[wiersz],
    )
    kreator = FakeKreator(stan)
    kontrolka = krok2_dane_kampanii.buduj(kreator)
    przycisk = _znajdz_przez_tekst(kontrolka, ft.FilledButton, "Wczytaj wiersze")
    przycisk.on_click(None)

    assert stan.uwagi == ""


def test_krok2_formularz_dalej_uzywa_aktualnego_stanu_a_nie_zamrozonego():
    """Regresja: 'Dalej' kiedyś liczył brakujące pola raz, przy pierwszym
    renderze (kiedy formularz jest pusty) — więc nawet po wypełnieniu
    wszystkiego użytkownik dostawał te same, nieaktualne błędy i utykał na
    kroku 2. Pola tekstowe/kombo celowo nie przebudowują widoku przy każdej
    zmianie (żeby nie tracić fokusu), więc walidacja przy kliknięciu musi
    czytać stan na bieżąco."""
    stan = StanKreatora(account_manager="Igor Samul", podmiot_realizujacy="Sp. k.", tryb_danych="formularz")
    kreator = FakeKreator(stan)

    kontrolka = krok2_dane_kampanii.buduj(kreator)  # zbudowany na PUSTYM stanie - rezerwuje nr_zlecenia

    for pole, wartosc in PELNY_STAN.items():
        # nr_zlecenia zostaje tym faktycznie zarezerwowanym powyżej - inaczej
        # rozjazd z nr_zlecenia_automatyczny wywołałby dialog ostrzeżenia o
        # numerze, co nie jest tym, co ten test sprawdza (patrz osobne testy
        # test_krok2_formularz_dalej_numer_*).
        if pole in ("account_manager", "podmiot_realizujacy", "nr_zlecenia"):
            continue
        setattr(stan, pole, wartosc)

    przycisk = _znajdz_przez_tekst(kontrolka, ft.FilledButton, "Dalej")
    assert przycisk is not None
    przycisk.on_click(None)

    assert kreator.bledy_pokazane == []
    assert stan.krok == 3


def test_krok2_formularz_dalej_bez_brandu_pokazuje_ostrzezenie_zamiast_przejsc():
    # Brand nie jest w pliku z kampaniami - łatwo o nim zapomnieć przy
    # wklejaniu wierszy, a bywa wymagany na zleceniu - stąd ostrzeżenie
    # zamiast cichego przejścia dalej.
    stan_kwargs = {k: v for k, v in PELNY_STAN.items() if k != "brand"}
    stan = StanKreatora(krok=2, tryb_danych="formularz", brand="", **stan_kwargs)
    kreator = FakeKreator(stan)
    kontrolka = krok2_dane_kampanii.buduj(kreator)

    _znajdz_przez_tekst(kontrolka, ft.FilledButton, "Dalej").on_click(None)

    assert stan.krok == 2  # nie przeskoczyło mimo poprawnych pozostałych pól
    assert kreator.page.dialog is not None
    assert "Brand" in kreator.page.dialog.content.value

    przycisk_kontynuuj = _znajdz_przez_tekst(kreator.page.dialog, ft.FilledButton, "Kontynuuj mimo to")
    assert przycisk_kontynuuj is not None
    przycisk_kontynuuj.on_click(None)
    assert stan.krok == 3


def test_krok2_formularz_dalej_z_brandem_nie_pokazuje_ostrzezenia():
    stan = StanKreatora(tryb_danych="formularz", **PELNY_STAN)
    kreator = FakeKreator(stan)
    kontrolka = krok2_dane_kampanii.buduj(kreator)

    _znajdz_przez_tekst(kontrolka, ft.FilledButton, "Dalej").on_click(None)

    assert kreator.page.dialog is None
    assert stan.krok == 3


def _stan_z_zarezerwowanym_numerem_i_reszta_wypelniona() -> tuple[StanKreatora, "FakeKreator", ft.Control]:
    stan = StanKreatora(krok=2, account_manager="Igor Samul", podmiot_realizujacy="Sp. k.", tryb_danych="formularz")
    kreator = FakeKreator(stan)
    kontrolka = krok2_dane_kampanii.buduj(kreator)  # rezerwuje nr_zlecenia_automatyczny naprawdę
    for pole, wartosc in PELNY_STAN.items():
        if pole in ("account_manager", "podmiot_realizujacy", "nr_zlecenia"):
            continue
        setattr(stan, pole, wartosc)
    return stan, kreator, kontrolka


def test_krok2_formularz_dalej_numer_rozny_od_automatycznego_pokazuje_ostrzezenie():
    stan, kreator, kontrolka = _stan_z_zarezerwowanym_numerem_i_reszta_wypelniona()
    numer_auto = stan.nr_zlecenia_automatyczny
    assert numer_auto is not None
    stan.nr_zlecenia = "K/2026/999"  # inny niż automatycznie pobrany

    _znajdz_przez_tekst(kontrolka, ft.FilledButton, "Dalej").on_click(None)

    assert stan.krok == 2  # nie przeskoczyło - czeka na decyzję
    assert kreator.page.dialog is not None
    tresc = kreator.page.dialog.content.value
    assert numer_auto in tresc
    assert "K/2026/999" in tresc


def test_krok2_formularz_dalej_numer_ten_sam_nie_pokazuje_ostrzezenia():
    stan, kreator, kontrolka = _stan_z_zarezerwowanym_numerem_i_reszta_wypelniona()
    # nr_zlecenia zostaje niezmienione (== nr_zlecenia_automatyczny)

    _znajdz_przez_tekst(kontrolka, ft.FilledButton, "Dalej").on_click(None)

    assert kreator.page.dialog is None
    assert stan.krok == 3


def test_krok2_formularz_numer_popraw_na_automatyczny_przywraca_numer():
    stan, kreator, kontrolka = _stan_z_zarezerwowanym_numerem_i_reszta_wypelniona()
    numer_auto = stan.nr_zlecenia_automatyczny
    stan.nr_zlecenia = "K/2026/999"

    _znajdz_przez_tekst(kontrolka, ft.FilledButton, "Dalej").on_click(None)
    przycisk = _znajdz_przez_tekst(kreator.page.dialog, ft.TextButton, f"Popraw na automatyczny ({numer_auto})")
    assert przycisk is not None
    przycisk.on_click(None)

    assert stan.nr_zlecenia == numer_auto
    assert stan.krok == 3
    assert kreator.page.dialog is None


def test_krok2_formularz_numer_zachowaj_wpisany_zwalnia_automatyczny_w_pliku():
    stan, kreator, kontrolka = _stan_z_zarezerwowanym_numerem_i_reszta_wypelniona()
    numer_auto = stan.nr_zlecenia_automatyczny
    stan.nr_zlecenia = "K/2026/999"

    _znajdz_przez_tekst(kontrolka, ft.FilledButton, "Dalej").on_click(None)
    przycisk = _znajdz_przez_tekst(kreator.page.dialog, ft.FilledButton, "Zachowaj wpisany (K/2026/999)")
    assert przycisk is not None
    przycisk.on_click(None)

    assert stan.nr_zlecenia == "K/2026/999"
    assert stan.krok == 3
    assert kreator.page.dialog is None

    # Rezerwacja automatycznego numeru w pliku ma zostać wyczyszczona.
    sciezka = ustawienia.wczytaj()["sciezka_numery_zlecen"]
    ws = openpyxl.load_workbook(sciezka)["Numery zleceń"]
    for r in range(2, ws.max_row + 1):
        if ws[f"A{r}"].value == numer_auto:
            assert ws[f"B{r}"].value is None
            assert ws[f"C{r}"].value is None
            assert ws[f"D{r}"].value is None
            break
    else:
        raise AssertionError(f"nie znaleziono wiersza numeru {numer_auto}")


def test_krok2_wklej_bez_konfliktu_przechodzi_do_formularza_do_uzupelnienia_brandu():
    """Po udanym wklejeniu wierszy apka pokazuje formularz (z już
    wypełnionymi polami) zamiast od razu skakać do okresów — Brand i Capping
    nie są częścią wklejanego wiersza i użytkownik musi mieć szansę je
    uzupełnić/sprawdzić."""
    stan = StanKreatora(krok=2, account_manager="Igor Samul", podmiot_realizujacy="Sp. k.", tryb_danych="wklej")
    kreator = FakeKreator(stan)
    krok2_dane_kampanii.buduj(kreator)

    wiersz = "\t".join(
        [
            "Test", "Initiative Media Warszawa sp. z o.o.", "Colian", "Paulina Kowalik",
            "KIDS", "NIE", "In-game audio KIDS", "Sp. k.", "", "CPM", "26",
            "K/2026/078", "1000", "01.07.2026", "31.07.2026",
        ]
    )
    stan.wiersze_wklejane = [wiersz]
    kontrolka = krok2_dane_kampanii.buduj(kreator)
    przycisk = _znajdz_przez_tekst(kontrolka, ft.FilledButton, "Wczytaj wiersze")
    przycisk.on_click(None)

    assert stan.krok == 2  # nie przeskoczyło do 3
    assert stan.tryb_danych == "formularz"


def test_krok2_plus_dodaje_kolejne_pole_wiersza():
    stan = StanKreatora(krok=2, account_manager="Igor Samul", podmiot_realizujacy="Sp. k.", tryb_danych="wklej")
    kreator = FakeKreator(stan)
    kontrolka = krok2_dane_kampanii.buduj(kreator)

    przycisk = _znajdz_przez_tekst(kontrolka, ft.OutlinedButton, "+ Dodaj wiersz (kolejny miesiąc)")
    assert przycisk is not None
    przycisk.on_click(None)

    assert stan.wiersze_wklejane == ["", ""]


def test_krok2_usun_wiersz_wycofuje_dodane_pole():
    stan = StanKreatora(
        krok=2, account_manager="Igor Samul", podmiot_realizujacy="Sp. k.", tryb_danych="wklej",
        wiersze_wklejane=["pierwszy", "drugi"],
    )
    kreator = FakeKreator(stan)
    kontrolka = krok2_dane_kampanii.buduj(kreator)

    przyciski_usun = _znajdz_wszystkie_ikony(kontrolka, ft.Icons.DELETE_OUTLINE)
    assert len(przyciski_usun) == 2
    przyciski_usun[1].on_click(None)  # usuń drugi wiersz

    assert stan.wiersze_wklejane == ["pierwszy"]


def test_krok2_plus_znika_po_osiagnieciu_limitu():
    stan = StanKreatora(
        krok=2, account_manager="Igor Samul", podmiot_realizujacy="Sp. k.", tryb_danych="wklej",
        wiersze_wklejane=[""] * krok2_dane_kampanii.LIMIT_WIERSZY_WKLEJANIA,
    )
    kreator = FakeKreator(stan)
    kontrolka = krok2_dane_kampanii.buduj(kreator)

    przycisk = _znajdz_przez_tekst(kontrolka, ft.OutlinedButton, "+ Dodaj wiersz (kolejny miesiąc)")
    assert przycisk is None


def test_krok2_wklej_dwa_osobne_pola_daje_dwa_okresy_tej_samej_kampanii():
    """Dwa miesiące tej samej kampanii przejściowej wklejone do dwóch osobnych
    pól (tak jak wklejone z dwóch różnych zakładek pliku kampanii) muszą się
    połączyć w jedną kampanię z dwoma okresami - bez konfliktu, bo pola
    wspólne (klient, format, model...) są takie same."""
    wiersz_lipiec = "\t".join(
        [
            "Test", "Initiative Media Warszawa sp. z o.o.", "Colian", "Paulina Kowalik",
            "KIDS", "TAK", "In-game audio KIDS", "Sp. k.", "", "CPM", "26",
            "K/2026/078", "1000", "01.07.2026", "31.07.2026",
        ]
    )
    wiersz_sierpien = "\t".join(
        [
            "Test", "Initiative Media Warszawa sp. z o.o.", "Colian", "Paulina Kowalik",
            "KIDS", "TAK", "In-game audio KIDS", "Sp. k.", "", "CPM", "26",
            "K/2026/078", "2000", "01.08.2026", "31.08.2026",
        ]
    )
    stan = StanKreatora(
        krok=2, account_manager="Igor Samul", podmiot_realizujacy="Sp. k.", tryb_danych="wklej",
        wiersze_wklejane=[wiersz_lipiec, wiersz_sierpien],
    )
    kreator = FakeKreator(stan)
    kontrolka = krok2_dane_kampanii.buduj(kreator)
    przycisk = _znajdz_przez_tekst(kontrolka, ft.FilledButton, "Wczytaj wiersze")
    przycisk.on_click(None)

    # 2+ okresy zawsze pokazują podsumowanie połączenia do potwierdzenia
    assert kreator.page.dialog is not None
    przycisk_potwierdz = _znajdz_przez_tekst(kreator.page.dialog, ft.FilledButton, "Potwierdź i kontynuuj")
    assert przycisk_potwierdz is not None
    przycisk_potwierdz.on_click(None)

    assert kreator.bledy_pokazane == []
    assert len(stan.okresy) == 2
    assert round(sum(o.budzet for o in stan.okresy), 2) == 3000.0
    assert stan.klient == "Colian"
    assert stan.okresy and stan.okresy[0].budzet == 1000.0


def _zawiera_etykiete(controls, etykieta) -> bool:
    for c in controls:
        if isinstance(c, ft.Text) and c.value == etykieta:
            return True
        content = getattr(c, "content", None)
        if isinstance(content, ft.Control) and _zawiera_etykiete([content], etykieta):
            return True
        nested = getattr(c, "controls", None) or []
        if nested and _zawiera_etykiete(nested, etykieta):
            return True
    return False


def _znajdz_pole_tekstowe_po_etykiecie(control, etykieta):
    """Szuka TextField albo po jego label=, albo po wzorcu ze Stack, gdzie
    TextField i Text(etykieta) (wcięty w ramkę - patrz
    pole_z_etykieta_na_ramce w pola_pomocnicze.py) są rodzeństwem w tym
    samym kontenerze, zamiast TextField.label, żeby etykieta nie wjeżdżała
    w środek pustego pola jak wpisana wartość."""
    if isinstance(control, ft.TextField) and control.label == etykieta:
        return control
    controls = getattr(control, "controls", None) or []
    pola = [c for c in controls if isinstance(c, ft.TextField)]
    if pola and _zawiera_etykiete(controls, etykieta):
        return pola[0]
    for c in controls:
        wynik = _znajdz_pole_tekstowe_po_etykiecie(c, etykieta)
        if wynik:
            return wynik
    content = getattr(control, "content", None)
    if isinstance(content, ft.Control):
        return _znajdz_pole_tekstowe_po_etykiecie(content, etykieta)
    return None


def test_krok3_dodaj_jeden_okres_odrzuca_zakres_na_dwa_miesiace():
    # "Dodaj jako jeden okres" musi odrzucić zakres rozciągnięty na dwa
    # miesiące — jeden okres = jeden wiersz w jednej zakładce miesięcznej
    # pliku kampanii, dla dwóch miesięcy trzeba użyć "Rozbij automatycznie".
    stan = StanKreatora(
        krok=3, account_manager="Igor Samul", podmiot_realizujacy="Sp. k.",
        nowy_okres_start=date(2026, 8, 15), nowy_okres_koniec=date(2026, 9, 15),
        nowy_okres_budzet="1000",
    )
    kreator = FakeKreator(stan)
    kontrolka = krok3_okresy.buduj(kreator)
    przycisk = _znajdz_przez_tekst(kontrolka, ft.OutlinedButton, "Dodaj jako jeden okres")
    przycisk.on_click(None)
    assert stan.okresy == []
    assert kreator.bledy_pokazane != []
    assert any("jednym miesiącu" in b for b in kreator.bledy_pokazane[0])


def test_krok3_rozbij_automatycznie_akceptuje_zakres_na_dwa_miesiace():
    # Regresja: naprawa powyższego (blokada wielomiesięcznego "jednego
    # okresu") nie może zablokować "Rozbij automatycznie" - to jego jedyne
    # zadanie, dzielić właśnie taki zakres na miesięczne kawałki.
    stan = StanKreatora(
        krok=3, account_manager="Igor Samul", podmiot_realizujacy="Sp. k.",
        nowy_okres_start=date(2026, 8, 15), nowy_okres_koniec=date(2026, 9, 15),
        nowy_okres_budzet="1000",
    )
    kreator = FakeKreator(stan)
    kontrolka = krok3_okresy.buduj(kreator)
    przycisk = _znajdz_przez_tekst(
        kontrolka, ft.OutlinedButton, "Rozbij automatycznie na miesiące (proporcjonalnie do dni)"
    )
    przycisk.on_click(None)
    assert len(stan.okresy) == 2
    assert kreator.bledy_pokazane == []


def test_krok3_pole_kosztu_wylacza_sie_przy_modelu_ff():
    """Regresja: przy modelu FF nie ma sensu koszt jednostkowy (cały budżet
    to jedna opłata) — pole ma być wyszarzone/nieedytowalne, nie wymagane."""
    stan = StanKreatora(
        krok=3, account_manager="Igor Samul", podmiot_realizujacy="Sp. k.", model_sprzedazy="FF",
    )
    kontrolka = krok3_okresy.buduj(FakeKreator(stan))
    pole = _znajdz_pole_tekstowe_po_etykiecie(kontrolka, "FF")
    assert pole is not None
    assert pole.disabled is True


def test_krok3_pole_kosztu_aktywne_dla_cpm():
    stan = StanKreatora(
        krok=3, account_manager="Igor Samul", podmiot_realizujacy="Sp. k.", model_sprzedazy="CPM",
    )
    kontrolka = krok3_okresy.buduj(FakeKreator(stan))
    pole = _znajdz_pole_tekstowe_po_etykiecie(kontrolka, "CPM")
    assert pole is not None
    assert pole.disabled is False


def test_krok3_dalej_wymaga_kosztu_dla_cpm_ale_nie_dla_ff():
    stan_cpm = StanKreatora(
        krok=3, account_manager="Igor Samul", podmiot_realizujacy="Sp. k.", model_sprzedazy="CPM",
        koszt_jednostkowy="", okresy=[Okres(date(2026, 7, 1), date(2026, 7, 31), 1000.0)],
    )
    kreator_cpm = FakeKreator(stan_cpm)
    kontrolka = krok3_okresy.buduj(kreator_cpm)
    _znajdz_przez_tekst(kontrolka, ft.FilledButton, "Dalej").on_click(None)
    assert kreator_cpm.bledy_pokazane != []
    assert stan_cpm.krok == 3

    stan_ff = StanKreatora(
        krok=3, account_manager="Igor Samul", podmiot_realizujacy="Sp. k.", model_sprzedazy="FF",
        koszt_jednostkowy="", okresy=[Okres(date(2026, 7, 1), date(2026, 7, 31), 1000.0)],
    )
    kreator_ff = FakeKreator(stan_ff)
    kontrolka = krok3_okresy.buduj(kreator_ff)
    _znajdz_przez_tekst(kontrolka, ft.FilledButton, "Dalej").on_click(None)
    assert kreator_ff.bledy_pokazane == []
    assert stan_ff.krok == 4
