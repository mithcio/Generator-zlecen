import json
from datetime import date, datetime

import openpyxl
import pytest

from app.models.kampania import PolaWspolne
from app.models.okres import Okres
from app.models.podmiot import SpolkaMediafarm
from app.models.zlecenie import Zlecenie
from app.services import cennik, generator_wydawcy as gw


def _zlecenie(format_reklamowy="Interstitial KIDS", podmiot_realizujacy="Sp. k.", target="KIDS", capping=3, uwagi="Dzieci"):
    pola = PolaWspolne(
        account_manager="Igor Samul",
        podmiot_realizujacy=podmiot_realizujacy,
        nr_zlecenia="K/2026/077",
        nazwa_kampanii="Colian_Hellena",
        dom_mediowy="Initiative Media Warszawa sp. z o.o.",
        klient="Colian",
        brand="Hellena",
        zlecajacy="Paulina Kowalik",
        target=target,
        capping=capping,
        format_reklamowy=format_reklamowy,
        model_sprzedazy="CPM",
        koszt_jednostkowy=26,
        uwagi=uwagi,
    )
    okresy = [Okres(date(2026, 7, 1), date(2026, 7, 31), 5000.0)]
    return Zlecenie(pola=pola, okresy=okresy)


SPOLKA = SpolkaMediafarm(
    nazwa="Mediafarm sp. k. sp. z o.o.",
    numery_rejestrowe="KRS: 0000708671 NIP: 527-283-08-24",
    konto_bankowe="mBank 38 1140 2004 0000 3102 7730 8080",
    adres="ul. Grzybowska 80/82, 00-844 Warszawa",
)
KONTAKT = {"email": "igor.samul@mediafarm.pl", "telefon": "500099699"}

# Odpowiada rzeczywistej strukturze zakładki "Traffic cennik" (Wydawca /
# Format / Stawka) - format to nazwa placementu, nie dokładny
# format_reklamowy z reszty aplikacji.
CENNIK_TESTOWY = {
    "KIDOZ": {
        "Interstitial": {"cena": 2.5, "waluta": "USD"},
        "Rewarded": {"cena": 3.5, "waluta": "USD"},
    },
    "PRADO": {
        "Interstitial": {"cena": 2.0, "waluta": "USD"},
        "Rewarded": {"cena": 2.0, "waluta": "USD"},
    },
    "Adverty": {
        "Display": {"cena": 3.5, "waluta": "USD"},
        "Video": {"cena": 5.0, "waluta": "USD"},
    },
    "Odeeo": {
        "In-Game Audio": {"cena": 2.0, "waluta": "EUR"},
    },
    "Crazygames": {
        "Video": {"cena": 2.5, "waluta": "EUR"},
    },
    "POKI": {
        "ImViTa": {"cena": 10.0, "waluta": "EUR"},
        "HPTO": {"cena": 7.0, "waluta": "EUR"},
        "Overlay": {"cena": 2.5, "waluta": "EUR"},
        "Rewarded": {"cena": 3.0, "waluta": "EUR"},
    },
}


@pytest.fixture(autouse=True)
def _cennik_testowy(tmp_path, monkeypatch):
    plik = tmp_path / "cennik_wydawcow.json"
    plik.write_text(json.dumps(CENNIK_TESTOWY), encoding="utf-8")
    monkeypatch.setattr(cennik, "DATA_PLIK", plik)


def test_generuj_kidoz_wypelnia_pola(tmp_path):
    zlecenie = _zlecenie(format_reklamowy="Interstitial KIDS", podmiot_realizujacy="Sp. k.", capping=5)
    sciezka = gw.generuj_kidoz(zlecenie, SPOLKA, KONTAKT, tmp_path)
    assert sciezka.exists()
    assert sciezka.name == "KIDOZ_Purchase_K-2026-077_Interstitial_KIDS_Colian_Hellena.xlsx"

    ws = openpyxl.load_workbook(sciezka).active
    assert ws["D5"].value == "Mediafarm"
    assert ws["D6"].value == "Igor Samul"
    assert ws["D7"].value == "igor.samul@mediafarm.pl"
    assert ws["D8"].value == "Colian"
    assert ws["D9"].value == "USD"
    assert ws["D10"].value == "K/2026/077"
    assert ws["C14"].value == "Colian_Hellena"
    assert ws["D14"].value == datetime(2026, 7, 1)
    assert ws["E14"].value == datetime(2026, 7, 31)
    assert ws["G14"].value == "Interstitial"
    assert ws["I14"].value == round(zlecenie.liczba_total)
    assert ws["J14"].value == 2.5
    assert ws["N14"].value == "KIDS"  # Targeting - z formatu, nigdy z Uwagi
    assert ws["O14"].value == "Capp 5"
    # Formuły wewnętrzne (odwołujące się tylko do komórek w tym samym
    # arkuszu) mają zostać nietknięte.
    assert str(ws["K14"].value).startswith("=")
    assert str(ws["L14"].value).startswith("=")


def test_generuj_kidoz_rewarded_dla_formatow_nie_display(tmp_path):
    zlecenie = _zlecenie(format_reklamowy="Rewarded ADULTS", target="KIDS")
    ws = openpyxl.load_workbook(gw.generuj_kidoz(zlecenie, SPOLKA, KONTAKT, tmp_path)).active
    assert ws["G14"].value == "Rewarded"
    assert ws["J14"].value == 3.5  # KIDOZ (target KIDS), nie PRADO


def test_generuj_kidoz_mobistitial_liczy_sie_jako_interstitial(tmp_path):
    # Rozszerzenie względem oryginalnej formuły szablonu - ta nie uwzględniała
    # Mobistitial w ogóle (wpadałoby w "Rewarded").
    zlecenie = _zlecenie(format_reklamowy="Mobistitial ADULTS", target="KIDS")
    ws = openpyxl.load_workbook(gw.generuj_kidoz(zlecenie, SPOLKA, KONTAKT, tmp_path)).active
    assert ws["G14"].value == "Interstitial"
    assert ws["J14"].value == 2.5


def test_generuj_kidoz_agencja_dla_spzoo(tmp_path):
    zlecenie = _zlecenie(podmiot_realizujacy="Sp. z o.o.")
    ws = openpyxl.load_workbook(gw.generuj_kidoz(zlecenie, SPOLKA, KONTAKT, tmp_path)).active
    assert ws["D5"].value == "Mediafarm Spółka"


def test_generuj_kidoz_bez_cappingu(tmp_path):
    zlecenie = _zlecenie(capping=None)
    ws = openpyxl.load_workbook(gw.generuj_kidoz(zlecenie, SPOLKA, KONTAKT, tmp_path)).active
    assert ws["O14"].value == "Capp brak"


def test_generuj_kidoz_brak_stawki_w_cenniku_podnosi_blad(tmp_path):
    cennik.DATA_PLIK.write_text(json.dumps({}), encoding="utf-8")
    with pytest.raises(cennik.BladCennika):
        gw.generuj_kidoz(_zlecenie(), SPOLKA, KONTAKT, tmp_path)


def test_kidoz_wydawca_cennika_target_kids():
    assert gw._kidoz_wydawca_cennika("KIDS") == "KIDOZ"


def test_kidoz_wydawca_cennika_target_adults_i_silver():
    assert gw._kidoz_wydawca_cennika("ADULTS") == "PRADO"
    assert gw._kidoz_wydawca_cennika("SILVER") == "PRADO"


def test_generuj_kidoz_target_adults_uzywa_cennika_prado(tmp_path):
    # To samo IO/plik co KIDOZ, ale inna stawka dla targetu Adults - patrz
    # ustalenia z użytkownikiem.
    zlecenie = _zlecenie(format_reklamowy="Rewarded ADULTS", target="ADULTS")
    ws = openpyxl.load_workbook(gw.generuj_kidoz(zlecenie, SPOLKA, KONTAKT, tmp_path)).active
    assert ws["J14"].value == 2.0  # PRADO Rewarded, nie KIDOZ Rewarded (3.5)
    assert ws["D5"].value == "Mediafarm"  # reszta pliku identyczna jak KIDOZ


def test_generuj_kidoz_target_silver_uzywa_cennika_prado(tmp_path):
    zlecenie = _zlecenie(target="SILVER")
    ws = openpyxl.load_workbook(gw.generuj_kidoz(zlecenie, SPOLKA, KONTAKT, tmp_path)).active
    assert ws["J14"].value == 2.0  # PRADO Interstitial, nie KIDOZ (2.5)


def test_generuj_adverty_display(tmp_path):
    zlecenie = _zlecenie(format_reklamowy="In-game DOOH DISPLAY ADULTS")
    sciezka = gw.generuj_adverty(zlecenie, SPOLKA, KONTAKT, tmp_path)
    assert sciezka.name == "Adverty_Purchase_K-2026-077_In-game_DOOH_DISPLAY_ADULTS_Colian_Hellena.xlsx"
    ws = openpyxl.load_workbook(sciezka).active

    assert ws["B8"].value == "Mediafarm sp. k. sp. z o.o."
    assert ws["B9"].value == "Initiative Media Warszawa sp. z o.o."
    assert ws["B10"].value == "Colian"
    assert ws["B11"].value == "Colian_Hellena"
    assert ws["B14"].value == round(zlecenie.liczba_total)
    assert ws["B15"].value == 3.5
    assert ws["B19"].value == "Y"
    assert ws["B20"].value == "N"
    assert str(ws["C19"].value).startswith("=")  # impression breakdown - formuła wewnętrzna
    assert ws["B30"].value == "ADULTS"  # Targeting Details - z formatu, nigdy z Uwagi


def test_generuj_adverty_video(tmp_path):
    zlecenie = _zlecenie(format_reklamowy="In-game DOOH VIDEO KIDS")
    ws = openpyxl.load_workbook(gw.generuj_adverty(zlecenie, SPOLKA, KONTAKT, tmp_path)).active
    assert ws["B19"].value == "N"
    assert ws["B20"].value == "Y"
    assert ws["B15"].value == 5.0


def test_generuj_odeeo_wypelnia_pola(tmp_path):
    zlecenie = _zlecenie(format_reklamowy="In-game audio KIDS")
    zlecenie.pola.nazwa_kampanii = "Colian_Hellena_Lody_Cooltowe_Lipiec_Sierpien_Audio"
    sciezka = gw.generuj_odeeo(zlecenie, SPOLKA, KONTAKT, tmp_path)
    assert sciezka.name == (
        "Odeeo_Purchase_K-2026-077_In-game_audio_KIDS_"
        "Colian_Hellena_Lody_Cooltowe_Lipiec_Sierpien_Audio.xlsx"
    )
    ws = openpyxl.load_workbook(sciezka).active

    assert ws["D3"].value == "Mediafarm sp. k. sp. z o.o."
    assert ws["D4"].value == "ul. Grzybowska 80/82, 00-844 Warszawa"
    assert ws["D5"].value == "+48 500 099 699"
    assert ws["D6"].value == "igor.samul@mediafarm.pl"
    assert ws["D7"].value == "Igor Samul"
    assert ws["A14"].value == "Colian_Hellena_Lody_Cooltowe_Lipiec_Sierpien_Audio_K/2026/077"
    assert ws["B14"].value == "01/07/2026 - 31/07/2026"
    assert ws["C14"].value == round(zlecenie.liczba_total)
    assert ws["D14"].value == 2.0
    assert ws["E17"].value == "Igor Samul"
    assert str(ws["E14"].value).startswith("=")  # Total Net Cost - formuła wewnętrzna
    assert str(ws["E16"].value).startswith("=")  # Company (=D3) - formuła wewnętrzna
    dzis = date.today()
    assert ws["E19"].value == datetime(dzis.year, dzis.month, dzis.day)  # data wystawienia
    assert ws["B6"].value is None  # kontakt Odeeo - zostawiony pusty
    assert ws["B7"].value is None
    assert ws.row_dimensions[14].height > 30.6  # rozciągnięte, żeby zmieściła się pełna nazwa


def test_generuj_odeeo_krotka_nazwa_nie_powieksza_wiersza_ponad_potrzebe(tmp_path):
    zlecenie = _zlecenie(format_reklamowy="In-game audio KIDS")
    zlecenie.pola.nazwa_kampanii = "X"
    ws = openpyxl.load_workbook(gw.generuj_odeeo(zlecenie, SPOLKA, KONTAKT, tmp_path)).active
    assert ws.row_dimensions[14].height == 30.6


def test_generuj_crazygames_wypelnia_pola(tmp_path):
    zlecenie = _zlecenie(capping=5)
    sciezka = gw.generuj_crazygames(zlecenie, SPOLKA, KONTAKT, tmp_path)
    assert sciezka.name == "Crazygames_Mediafarm_Colian_Colian_Hellena_IO.xlsx"
    ws = openpyxl.load_workbook(sciezka).active

    assert ws["A4"].value == "IO based campaign"  # niezmienione
    assert ws["B4"].value == "Colian"  # Advertiser = klient
    assert ws["C4"].value == datetime(2026, 7, 1)
    assert ws["D4"].value == datetime(2026, 7, 31)
    assert ws["E4"].value == "Video Pre-Roll"  # niezmienione
    assert ws["G4"].value == round(zlecenie.liczba_total)
    assert ws["I4"].value == 2.5  # z cennika (Crazygames, Video)
    assert ws["K4"].value == "5/campaign"
    assert str(ws["J4"].value).startswith("=")  # Budget - formuła wewnętrzna


def test_generuj_crazygames_bez_cappingu(tmp_path):
    zlecenie = _zlecenie(capping=None)
    ws = openpyxl.load_workbook(gw.generuj_crazygames(zlecenie, SPOLKA, KONTAKT, tmp_path)).active
    assert ws["K4"].value == "brak/campaign"


def test_generuj_poki_wypelnia_pola(tmp_path):
    zlecenie = _zlecenie(capping=3)
    sciezka = gw.generuj_poki(zlecenie, SPOLKA, KONTAKT, tmp_path, "ImViTa")
    assert sciezka.name == "POKI_Colian_Hellena_IO.xlsx"

    ws = openpyxl.load_workbook(sciezka).active
    wartosci = {ws.cell(row=r, column=1).value: ws.cell(row=r, column=2).value for r in range(1, 11)}

    assert wartosci["Campaign"] == "Colian_Hellena"
    assert wartosci["Timing"] == "01.07.2026-31.07.2026"
    assert wartosci["Format"] == "ImViTa"
    assert wartosci["Impressions"].endswith(" PVs")
    assert wartosci["Sites"] == "poki.pl"
    assert wartosci["Target"] == "Kids"
    assert wartosci["Geo"] == "Poland"
    assert wartosci["Capp"] == "3"
    assert wartosci["Device"] == "cross-device"
    assert wartosci["CPM"] == "10€"

    # kolumna A pogrubiona, kolumna B nie
    assert ws["A1"].font.bold is True
    assert ws["B1"].font.bold is not True


def test_generuj_poki_szerokosc_kolumn_dopasowana(tmp_path):
    ws = openpyxl.load_workbook(
        gw.generuj_poki(_zlecenie(), SPOLKA, KONTAKT, tmp_path, "Rewarded")
    ).active
    assert ws.column_dimensions["A"].width == len("Impressions") + 2  # najdłuższa etykieta
    assert ws.column_dimensions["B"].width > 0


def test_generuj_poki_brak_stawki_dla_placementu_podnosi_blad(tmp_path):
    with pytest.raises(cennik.BladCennika):
        gw.generuj_poki(_zlecenie(), SPOLKA, KONTAKT, tmp_path, "Nieznany placement")


def test_brak_szablonu_podnosi_czytelny_blad(tmp_path, monkeypatch):
    monkeypatch.setattr(gw, "SZABLONY_DIR", tmp_path / "nieistniejace")
    zlecenie = _zlecenie()
    with pytest.raises(gw.BladSzablonuWydawcy):
        gw.generuj_kidoz(zlecenie, SPOLKA, KONTAKT, tmp_path)


@pytest.mark.parametrize(
    "wydawca, format_reklamowy, oczekiwane",
    [
        ("KIDOZ", "Rewarded KIDS", True),
        ("KIDOZ", "Interstitial ADULTS", True),
        ("KIDOZ", "Mobistitial KIDS", True),
        ("KIDOZ", "In-game DOOH VIDEO KIDS", False),
        ("KIDOZ", "In-game audio KIDS", False),
        ("POKI", "Rewarded ADULTS", True),
        ("POKI", "Non-standard", True),
        ("POKI", "Interstitial KIDS", False),
        ("CRAZYGAMES", "Rewarded KIDS", True),
        ("CRAZYGAMES", "In-game audio KIDS", False),
        ("ADVERTY", "In-game DOOH DISPLAY ADULTS", True),
        ("ADVERTY", "In-game DOOH VIDEO KIDS", True),
        ("ADVERTY", "Rewarded KIDS", False),
        ("ODEEO", "In-game audio KIDS", True),
        ("ODEEO", "Rewarded KIDS", False),
        ("KIDOZ", "YouTube", False),
        ("POKI", "Video interaktywne", False),
    ],
)
def test_czy_format_pasuje(wydawca, format_reklamowy, oczekiwane):
    assert gw.czy_format_pasuje(wydawca, format_reklamowy) is oczekiwane


@pytest.mark.parametrize(
    "format_reklamowy, oczekiwany",
    [
        ("Interstitial KIDS", "KIDS"),
        ("Rewarded ADULTS", "ADULTS"),
        ("In-game DOOH DISPLAY KIDS", "KIDS"),
        ("YouTube", ""),
        ("Non-standard", ""),
    ],
)
def test_target_z_formatu(format_reklamowy, oczekiwany):
    assert gw._target_z_formatu(format_reklamowy) == oczekiwany


def test_generatory_dostepne_dla_glownych_wydawcow():
    # POKI nie jest w GENERATORY - potrzebuje dodatkowego argumentu
    # (placement), wywoływane wprost (patrz krok5_dane_traffic.py).
    assert set(gw.GENERATORY.keys()) == {"KIDOZ", "ADVERTY", "ODEEO", "CRAZYGAMES"}
