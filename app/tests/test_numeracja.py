import openpyxl
import pytest

from app.services import numeracja, ustawienia


def _zbuduj_plik_numery(tmp_path, ile=10, zajete_k=0, zajete_s=0):
    """Plik xlsx z tą samą strukturą co Numery_zlecen_2026.xlsx: arkusz
    "Numery zleceń", A-D = SP. K. (numer/status/data/użytkownik), F-I = SP. Z
    O.O. (to samo). `zajete_k`/`zajete_s` = ile pierwszych wierszy oznaczyć
    jako już zajęte (do testowania "pierwszy wolny wiersz")."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Numery zleceń"
    ws.append(["SP. K.", "status", "data", "użytkownik", None, "SP. Z O.O.", "status", "data", "użytkownik"])
    for i in range(1, ile + 1):
        wiersz = [f"K/2026/{i:03d}", "zajete" if i <= zajete_k else None, None, None, None,
                  f"S/2026/{i:03d}", "zajete" if i <= zajete_s else None, None, None]
        ws.append(wiersz)
    plik = tmp_path / "Numery_zlecen_2026.xlsx"
    wb.save(plik)
    return plik


@pytest.fixture(autouse=True)
def _izolacja_ustawien(tmp_path, monkeypatch):
    monkeypatch.setattr(ustawienia, "DATA_PLIK", tmp_path / "ustawienia.json")


def _ustaw_sciezke(plik):
    ustawienia.zapisz(sciezka_numery_zlecen=str(plik))


def test_zarezerwuj_numer_bierze_pierwszy_wolny_wiersz(tmp_path):
    plik = _zbuduj_plik_numery(tmp_path, ile=10, zajete_k=3)
    _ustaw_sciezke(plik)
    numer = numeracja.zarezerwuj_numer("Sp. k.")
    assert numer == "K/2026/004"


def test_zarezerwuj_numer_kolejny_raz_daje_kolejny_numer(tmp_path):
    plik = _zbuduj_plik_numery(tmp_path, ile=10)
    _ustaw_sciezke(plik)
    numeracja.zarezerwuj_numer("Sp. k.")
    drugi = numeracja.zarezerwuj_numer("Sp. k.")
    assert drugi == "K/2026/002"


def test_zarezerwuj_numer_zapisuje_status_date_i_akanta_w_pliku(tmp_path):
    plik = _zbuduj_plik_numery(tmp_path, ile=5)
    _ustaw_sciezke(plik)
    numer = numeracja.zarezerwuj_numer("Sp. k.", "Marta Urbańska")
    assert numer == "K/2026/001"

    wb = openpyxl.load_workbook(plik)
    ws = wb["Numery zleceń"]
    assert ws["B2"].value == "zajete"
    assert ws["C2"].value
    assert ws["D2"].value == "Marta Urbańska"


def test_zarezerwuj_numer_wypelnione_komorki_wysrodkowane(tmp_path):
    plik = _zbuduj_plik_numery(tmp_path, ile=5)
    _ustaw_sciezke(plik)
    numeracja.zarezerwuj_numer("Sp. k.", "Marta Urbańska")

    wb = openpyxl.load_workbook(plik)
    ws = wb["Numery zleceń"]
    for komorka in (ws["B2"], ws["C2"], ws["D2"]):
        assert komorka.alignment.horizontal == "center"


def test_zarezerwuj_numer_dwa_prefiksy_niezalezne(tmp_path):
    plik = _zbuduj_plik_numery(tmp_path, ile=10, zajete_k=3, zajete_s=1)
    _ustaw_sciezke(plik)
    assert numeracja.zarezerwuj_numer("Sp. k.") == "K/2026/004"
    assert numeracja.zarezerwuj_numer("Sp. z o.o.") == "S/2026/002"

    wb = openpyxl.load_workbook(plik)
    ws = wb["Numery zleceń"]
    assert ws["B6"].value is None  # kolejny wolny K wiersz nietknięty
    assert ws["G4"].value is None  # kolejny wolny S wiersz nietknięty


def test_zarezerwuj_numer_bez_sciezki_w_ustawieniach_podnosi_blad(tmp_path):
    with pytest.raises(numeracja.BladNumeracji):
        numeracja.zarezerwuj_numer("Sp. k.")


def test_zarezerwuj_numer_brak_wolnych_numerow_podnosi_blad(tmp_path):
    plik = _zbuduj_plik_numery(tmp_path, ile=3, zajete_k=3)
    _ustaw_sciezke(plik)
    with pytest.raises(numeracja.BladNumeracji):
        numeracja.zarezerwuj_numer("Sp. k.")


def test_zarezerwuj_numer_plik_otwarty_gdzie_indziej_daje_czytelny_blad(tmp_path, monkeypatch):
    # Regresja: PermissionError (plik otwarty w Excelu / w trakcie synchro-
    # nizacji z chmurą) wcześniej wylatywał niezłapany i wywalał całą stronę
    # kreatora zamiast pokazać komunikat do poprawienia przez użytkownika.
    plik = _zbuduj_plik_numery(tmp_path, ile=5)
    _ustaw_sciezke(plik)

    def _zajety(*args, **kwargs):
        raise PermissionError(13, "Permission denied", str(plik))

    monkeypatch.setattr(numeracja.openpyxl, "load_workbook", _zajety)
    with pytest.raises(numeracja.BladNumeracji):
        numeracja.zarezerwuj_numer("Sp. k.")


def test_zwolnij_numer_czysci_status_date_i_akanta(tmp_path):
    plik = _zbuduj_plik_numery(tmp_path, ile=5)
    _ustaw_sciezke(plik)
    numer = numeracja.zarezerwuj_numer("Sp. k.", "Marta Urbańska")

    numeracja.zwolnij_numer(numer, "Sp. k.")

    wb = openpyxl.load_workbook(plik)
    ws = wb["Numery zleceń"]
    assert ws["B2"].value is None
    assert ws["C2"].value is None
    assert ws["D2"].value is None
    assert ws["A2"].value == numer  # sam numer w kolumnie A zostaje


def test_zwolnij_numer_sp_zoo_czysci_kolumny_g_h_i(tmp_path):
    plik = _zbuduj_plik_numery(tmp_path, ile=5)
    _ustaw_sciezke(plik)
    numer = numeracja.zarezerwuj_numer("Sp. z o.o.", "Igor Samul")

    numeracja.zwolnij_numer(numer, "Sp. z o.o.")

    wb = openpyxl.load_workbook(plik)
    ws = wb["Numery zleceń"]
    assert ws["G2"].value is None
    assert ws["H2"].value is None
    assert ws["I2"].value is None


def test_zwolnij_numer_pozwala_ponownie_zarezerwowac_ten_sam_numer(tmp_path):
    plik = _zbuduj_plik_numery(tmp_path, ile=5)
    _ustaw_sciezke(plik)
    pierwszy = numeracja.zarezerwuj_numer("Sp. k.")
    numeracja.zwolnij_numer(pierwszy, "Sp. k.")

    ponownie = numeracja.zarezerwuj_numer("Sp. k.")
    assert ponownie == pierwszy  # znów pierwszy wolny wiersz


def test_zwolnij_numer_nieznaleziony_podnosi_blad(tmp_path):
    plik = _zbuduj_plik_numery(tmp_path, ile=5)
    _ustaw_sciezke(plik)
    with pytest.raises(numeracja.BladNumeracji):
        numeracja.zwolnij_numer("K/2026/999", "Sp. k.")


def test_prefiksy_niezalezne():
    assert numeracja.PREFIKSY["Sp. k."] == "K"
    assert numeracja.PREFIKSY["Sp. z o.o."] == "S"
