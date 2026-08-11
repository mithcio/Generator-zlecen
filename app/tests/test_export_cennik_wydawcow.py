"""scripts/export_seed_data.py nie ma zwykle testów (offline, ręcznie
uruchamiany skrypt) - ale export_cennik_wydawcow ma nietrywialną heurystykę
(waluta odczytywana z formatowania komórki, nie z osobnej kolumny), więc
warto ją sprawdzić wprost - błąd tutaj cicho zepsułby ceny w wygenerowanych
plikach dla wydawców."""
import json

import openpyxl

from scripts.export_seed_data import _waluta_z_formatu, export_cennik_wydawcow


def _zbuduj_plik_cennika(tmp_path):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Traffic cennik"
    ws.append(["Wydawca", "Format", "Stawka", "Uwagi"])
    ws.append(["KIDOZ", "Rewarded", 3.5, None])
    ws["C2"].number_format = '"$"#,##0.00'
    ws.append(["Odeeo", "In-Game Audio", 2.0, None])
    ws["C3"].number_format = "€#,##0.00"
    ws.append(["POKI", "ImViTa", 10.0, "Cena może być inna, zależnie od ustaleń akanta z POKI"])
    ws["C4"].number_format = "€#,##0.00"
    plik = tmp_path / "Numery_zlecen_2026.xlsx"
    wb.save(plik)
    return plik


def test_export_cennik_wydawcow_czyta_wydawce_format_i_stawke(tmp_path, monkeypatch):
    plik = _zbuduj_plik_cennika(tmp_path)
    monkeypatch.setattr("scripts.export_seed_data.DATA_OUT", tmp_path)
    export_cennik_wydawcow(plik)

    wynik = json.loads((tmp_path / "cennik_wydawcow.json").read_text(encoding="utf-8"))
    assert wynik["KIDOZ"]["Rewarded"] == {"cena": 3.5, "waluta": "USD"}
    assert wynik["Odeeo"]["In-Game Audio"] == {"cena": 2.0, "waluta": "EUR"}
    assert wynik["POKI"]["ImViTa"] == {"cena": 10.0, "waluta": "EUR"}


def test_export_cennik_wydawcow_brak_zakladki_daje_pusty_cennik(tmp_path, monkeypatch):
    wb = openpyxl.Workbook()
    plik = tmp_path / "Numery_zlecen_2026.xlsx"
    wb.save(plik)
    monkeypatch.setattr("scripts.export_seed_data.DATA_OUT", tmp_path)
    export_cennik_wydawcow(plik)
    wynik = json.loads((tmp_path / "cennik_wydawcow.json").read_text(encoding="utf-8"))
    assert wynik == {}


def test_waluta_z_formatu_dolar():
    assert _waluta_z_formatu('"$"#,##0.00') == "USD"


def test_waluta_z_formatu_euro():
    assert _waluta_z_formatu("€#,##0.00") == "EUR"


def test_waluta_z_formatu_zloty():
    assert _waluta_z_formatu('#,##0.00" zł"') == "PLN"


def test_waluta_z_formatu_domyslnie_usd():
    assert _waluta_z_formatu("General") == "USD"
