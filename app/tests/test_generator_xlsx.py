import re
from datetime import date

import openpyxl

from app.models.kampania import PolaWspolne
from app.models.okres import Okres
from app.models.zlecenie import Zlecenie
from app.services import lookup_podmiotu as lp
from app.services.generator_xlsx import generuj_xlsx


def _przykladowe_zlecenie(model_sprzedazy="CPM", koszt_jednostkowy=26, capping=3):
    pola = PolaWspolne(
        account_manager="Igor Samul",
        podmiot_realizujacy="Sp. k.",
        nr_zlecenia="K/2026/077",
        nazwa_kampanii="Colian_Hellena_Lody_Cooltowe_Lipiec_Sierpien_Audio",
        dom_mediowy="Initiative Media Warszawa sp. z o.o.",
        klient="Colian",
        brand="Hellena",
        zlecajacy="Paulina Kowalik",
        target="KIDS",
        capping=capping,
        format_reklamowy="In-game audio KIDS",
        model_sprzedazy=model_sprzedazy,
        koszt_jednostkowy=koszt_jednostkowy,
        uwagi="Dzieci, Młodzież, Dorośli",
    )
    okresy = [
        Okres(date(2026, 7, 28), date(2026, 7, 31), 5714.29),
        Okres(date(2026, 8, 1), date(2026, 8, 31), 44285.71),
    ]
    return Zlecenie(pola=pola, okresy=okresy)


_WZORZEC_NUMERU_POZYCJI = re.compile(r"^\d+\.\d+$")


def _wczytaj_komorki_po_numerze(sciezka):
    """Mapuje numer pozycji ('4.5', '7.1', ...) na komórkę wartości (kolumna C).
    Dopasowuje tylko prawdziwe numery Pozycja ("N.N") - nie nagłówki sekcji
    takie jak "1. Dane Mediafarm", które też zaczynają się od cyfry."""
    wb = openpyxl.load_workbook(sciezka)
    ws = wb.active
    wynik = {}
    for row in ws.iter_rows():
        numer = row[0].value
        if isinstance(numer, str) and _WZORZEC_NUMERU_POZYCJI.match(numer):
            wynik[numer] = row[2]  # kolumna C
    return wynik


def test_generuj_xlsx_kluczowe_komorki_tekstowe(tmp_path):
    zlecenie = _przykladowe_zlecenie()
    podmiot = lp.znajdz_podmiot(zlecenie.pola.account_manager, zlecenie.pola.dom_mediowy)
    spolka = lp.spolka_mediafarm(zlecenie.pola.podmiot_realizujacy)
    kontakt = lp.kontakt_accounta(zlecenie.pola.account_manager)

    sciezka = generuj_xlsx(zlecenie, podmiot, spolka, kontakt, tmp_path / "Zlecenie_test.xlsx")
    assert sciezka.exists()

    wb = openpyxl.load_workbook(sciezka)
    ws = wb.active
    tekst_calego_arkusza = "\n".join(
        str(c.value) for row in ws.iter_rows() for c in row if c.value is not None
    )
    assert "K/2026/077" in tekst_calego_arkusza
    assert "Colian" in tekst_calego_arkusza
    assert "Hellena" in tekst_calego_arkusza
    assert "28.07.2026" in tekst_calego_arkusza
    assert "31.08.2026" in tekst_calego_arkusza
    # Nota o rachunku ma być osobnym polem, nie sklejona z kwotą 7.3.
    assert "Kwota brutto do zapłaty na rachunek wskazany w pkt. 1.2." in tekst_calego_arkusza


def test_liczba_i_brutto_sa_prawdziwymi_formulami_cpm(tmp_path):
    """4.5 i 7.3 muszą być formułami Excela odwołującymi się do 4.6/7.1/7.2,
    nie statycznym tekstem — żeby edycja budżetu netto w Excelu przeliczała
    resztę automatycznie."""
    zlecenie = _przykladowe_zlecenie(model_sprzedazy="CPM", koszt_jednostkowy=26)
    podmiot = lp.znajdz_podmiot(zlecenie.pola.account_manager, zlecenie.pola.dom_mediowy)
    spolka = lp.spolka_mediafarm(zlecenie.pola.podmiot_realizujacy)
    kontakt = lp.kontakt_accounta(zlecenie.pola.account_manager)

    sciezka = generuj_xlsx(zlecenie, podmiot, spolka, kontakt, tmp_path / "Zlecenie_test.xlsx")
    komorki = _wczytaj_komorki_po_numerze(sciezka)

    assert komorki["4.6"].value == 26  # koszt jednostkowy - prawdziwa liczba
    assert komorki["7.1"].value == 50000.0  # koszt netto - prawdziwa liczba
    assert komorki["7.2"].value == 0.23

    formula_liczby = komorki["4.5"].value
    assert formula_liczby.startswith("=")
    assert f"C{komorki['7.1'].row}" in formula_liczby
    assert f"C{komorki['4.6'].row}" in formula_liczby
    assert "*1000" in formula_liczby  # CPM

    formula_brutto = komorki["7.3"].value
    assert formula_brutto.startswith("=")
    assert f"C{komorki['7.1'].row}" in formula_brutto
    assert f"C{komorki['7.2'].row}" in formula_brutto


def test_liczba_formula_cpc_bez_mnoznika(tmp_path):
    zlecenie = _przykladowe_zlecenie(model_sprzedazy="CPC", koszt_jednostkowy=2)
    podmiot = lp.znajdz_podmiot(zlecenie.pola.account_manager, zlecenie.pola.dom_mediowy)
    spolka = lp.spolka_mediafarm(zlecenie.pola.podmiot_realizujacy)
    kontakt = lp.kontakt_accounta(zlecenie.pola.account_manager)

    sciezka = generuj_xlsx(zlecenie, podmiot, spolka, kontakt, tmp_path / "Zlecenie_test.xlsx")
    komorki = _wczytaj_komorki_po_numerze(sciezka)
    assert "*1000" not in komorki["4.5"].value


def test_ff_koszt_rowny_budzetowi_i_liczba_jeden(tmp_path):
    """FF (opłata stała) nie ma kosztu jednostkowego — 4.6 pokazuje budżet
    wprost, a 4.5 (liczba) zawsze wynosi 1."""
    zlecenie = _przykladowe_zlecenie(model_sprzedazy="FF", koszt_jednostkowy=0)
    podmiot = lp.znajdz_podmiot(zlecenie.pola.account_manager, zlecenie.pola.dom_mediowy)
    spolka = lp.spolka_mediafarm(zlecenie.pola.podmiot_realizujacy)
    kontakt = lp.kontakt_accounta(zlecenie.pola.account_manager)

    sciezka = generuj_xlsx(zlecenie, podmiot, spolka, kontakt, tmp_path / "Zlecenie_test.xlsx")
    komorki = _wczytaj_komorki_po_numerze(sciezka)

    assert komorki["4.6"].value == 50000.0  # = budżet netto
    assert komorki["4.5"].value == 1


def test_cala_kolumna_c_wyrownana_do_lewej(tmp_path):
    """Excel domyślnie wyrównuje liczby do prawej, a tekst do lewej — bez
    jawnego wyrównania kolumna C (mieszanka tekstu i prawdziwych liczb/formuł
    4.5/4.6/7.1-7.3) wyglądałaby na rozjechaną. Cała kolumna C ma być do lewej."""
    zlecenie = _przykladowe_zlecenie()
    podmiot = lp.znajdz_podmiot(zlecenie.pola.account_manager, zlecenie.pola.dom_mediowy)
    spolka = lp.spolka_mediafarm(zlecenie.pola.podmiot_realizujacy)
    kontakt = lp.kontakt_accounta(zlecenie.pola.account_manager)

    sciezka = generuj_xlsx(zlecenie, podmiot, spolka, kontakt, tmp_path / "Zlecenie_test.xlsx")
    komorki = _wczytaj_komorki_po_numerze(sciezka)

    for numer, komorka in komorki.items():
        assert komorka.alignment.horizontal == "left", f"{numer}: {komorka.alignment.horizontal!r}"


def test_capping_zapisany_jako_liczba(tmp_path):
    """Capping ma być prawdziwą liczbą w xlsx (nie tekstem), żeby dało się
    jej użyć w przyszłej formule — "brak" zostaje tekstem, bo nie ma sensownej
    wartości liczbowej."""
    zlecenie = _przykladowe_zlecenie(capping=7)
    podmiot = lp.znajdz_podmiot(zlecenie.pola.account_manager, zlecenie.pola.dom_mediowy)
    spolka = lp.spolka_mediafarm(zlecenie.pola.podmiot_realizujacy)
    kontakt = lp.kontakt_accounta(zlecenie.pola.account_manager)

    sciezka = generuj_xlsx(zlecenie, podmiot, spolka, kontakt, tmp_path / "Zlecenie_test.xlsx")
    komorki = _wczytaj_komorki_po_numerze(sciezka)
    assert komorki["4.3"].value == 7
    assert isinstance(komorki["4.3"].value, int)


def test_capping_brak_zostaje_tekstem(tmp_path):
    zlecenie = _przykladowe_zlecenie(capping=None)
    podmiot = lp.znajdz_podmiot(zlecenie.pola.account_manager, zlecenie.pola.dom_mediowy)
    spolka = lp.spolka_mediafarm(zlecenie.pola.podmiot_realizujacy)
    kontakt = lp.kontakt_accounta(zlecenie.pola.account_manager)

    sciezka = generuj_xlsx(zlecenie, podmiot, spolka, kontakt, tmp_path / "Zlecenie_test.xlsx")
    komorki = _wczytaj_komorki_po_numerze(sciezka)
    assert komorki["4.3"].value == "brak"
